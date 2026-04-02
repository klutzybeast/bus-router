"""Google Sheets integration and seat availability endpoints."""

import os
import csv
import io
import logging
from collections import defaultdict
from datetime import datetime, timezone
from typing import Optional

import httpx
from fastapi import APIRouter, HTTPException
from fastapi.responses import Response, StreamingResponse

from services.database import (
    db, sheets_generator, cover_sheet_generator, OUTPUT_SHEET_ID
)
from services.helpers import get_active_season_id
from bus_config import (
    get_bus_capacity, get_bus_location, get_bus_driver, get_bus_counselor
)

logger = logging.getLogger(__name__)

router = APIRouter(tags=["Sheets"])

@router.get("/sheets/seat-availability")
async def get_seat_availability_for_sheets():
    """Get formatted seat availability data for Google Sheets - COVER SHEET FORMAT"""
    try:
        # Include ALL campers with bus assignments (even without valid locations)
        campers = await db.campers.find({
            "am_bus_number": {"$exists": True, "$nin": ["NONE", ""]}
        }).to_list(None)
        
        # Get staff configurations from database
        staff_configs = await db.bus_staff.find({}).to_list(None)
        staff_dict = {c['bus_number']: c for c in staff_configs}
        
        # Get shadows, assigned staff, and staff with addresses for notes column
        shadows = await db.shadows.find({}).to_list(None)
        assigned_staff = await db.bus_assigned_staff.find({}).to_list(None)
        staff_addresses = await db.staff_addresses.find({}).to_list(None)
        
        # Use compact Cover Sheet format with staff info and notes
        sheet_data = cover_sheet_generator.generate_cover_sheet(campers, staff_dict, shadows, assigned_staff, staff_addresses)
        
        return {
            "status": "success",
            "data": sheet_data,
            "last_updated": datetime.now(timezone.utc).isoformat()
        }
    except Exception as e:
        logging.error(f"Error generating sheets data: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/sheets/compact-availability")
async def get_compact_availability():
    """Get compact seat availability summary for Google Sheets"""
    try:
        campers = await db.campers.find({
            "bus_number": {"$exists": True, "$nin": ["NONE", ""]}
        }).to_list(None)
        compact_data = sheets_generator.generate_compact_availability(campers)
        
        return {
            "status": "success",
            "data": compact_data,
            "last_updated": datetime.now(timezone.utc).isoformat()
        }
    except Exception as e:
        logging.error(f"Error generating compact data: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))
@router.get("/seat-availability-json")
async def get_seat_availability_json():
    """Get seat availability data as JSON for frontend display"""
    try:
        # Get ALL campers with bus assignments (including those without addresses)
        campers = await db.campers.find({
            "$or": [
                {"am_bus_number": {"$exists": True, "$nin": ["NONE", ""]}},
                {"pm_bus_number": {"$exists": True, "$nin": ["NONE", ""]}}
            ]
        }).to_list(None)
        
        # Get all shadows (they take bus seats too)
        shadows = await db.shadows.find({}).to_list(None)
        
        # Get all assigned staff (they also take bus seats)
        assigned_staff = await db.bus_assigned_staff.find({}).to_list(None)
        
        # Get all staff with addresses (they also take bus seats)
        staff_addresses = await db.staff_addresses.find({}).to_list(None)
        
        # Get staff configurations from database
        staff_configs = await db.bus_staff.find({}).to_list(None)
        staff_dict = {c['bus_number']: c for c in staff_configs}
        
        # Group and count by bus
        from collections import defaultdict
        bus_data = defaultdict(lambda: {
            'h1_am': 0, 'h1_pm': 0, 'h2_am': 0, 'h2_pm': 0,
            'shadows': 0, 'assigned_staff': 0, 'staff_with_addresses': 0,
            'capacity': 30, 'location': '', 'driver': 'TBD', 'counselor': 'TBD'
        })
        
        def parse_session(session):
            """Parse session to determine which halves the camper attends"""
            session_lower = (session or '').lower()
            is_full = 'full season' in session_lower or 'full' in session_lower
            is_half1 = 'half season 1' in session_lower or 'half 1' in session_lower or 'first half' in session_lower
            is_half2 = 'half season 2' in session_lower or 'half 2' in session_lower or 'second half' in session_lower
            is_flex = '6 week' in session_lower or 'flex' in session_lower
            
            # Default to full if no session specified
            if not is_full and not is_half1 and not is_half2 and not is_flex:
                is_full = True
            
            return {
                'h1': is_full or is_half1 or is_flex,
                'h2': is_full or is_half2 or is_flex
            }
        
        # Process each camper
        for camper in campers:
            am_bus = camper.get('am_bus_number', '')
            pm_bus = camper.get('pm_bus_number', '')
            session = camper.get('session', '')
            halves = parse_session(session)
            
            # Count for AM bus
            if am_bus and am_bus != 'NONE' and am_bus.startswith('Bus'):
                if halves['h1']:
                    bus_data[am_bus]['h1_am'] += 1
                if halves['h2']:
                    bus_data[am_bus]['h2_am'] += 1
            
            # Count for PM bus
            if pm_bus and pm_bus != 'NONE' and pm_bus.startswith('Bus'):
                if halves['h1']:
                    bus_data[pm_bus]['h1_pm'] += 1
                if halves['h2']:
                    bus_data[pm_bus]['h2_pm'] += 1
        
        # Process shadows - they inherit the session from their linked camper
        for shadow in shadows:
            bus_number = shadow.get('bus_number', '')
            session = shadow.get('session', '')
            halves = parse_session(session)
            
            if bus_number and bus_number.startswith('Bus'):
                bus_data[bus_number]['shadows'] += 1
                # Shadows take both AM and PM seats (same as their linked camper)
                if halves['h1']:
                    bus_data[bus_number]['h1_am'] += 1
                    bus_data[bus_number]['h1_pm'] += 1
                if halves['h2']:
                    bus_data[bus_number]['h2_am'] += 1
                    bus_data[bus_number]['h2_pm'] += 1
        
        # Process assigned staff - they take seats too
        for staff in assigned_staff:
            bus_number = staff.get('bus_number', '')
            session = staff.get('session', '')
            halves = parse_session(session)
            
            if bus_number and bus_number.startswith('Bus'):
                bus_data[bus_number]['assigned_staff'] += 1
                # Assigned staff take both AM and PM seats
                if halves['h1']:
                    bus_data[bus_number]['h1_am'] += 1
                    bus_data[bus_number]['h1_pm'] += 1
                if halves['h2']:
                    bus_data[bus_number]['h2_am'] += 1
                    bus_data[bus_number]['h2_pm'] += 1
        
        # Process staff with addresses - they take seats too (only if bus assigned)
        for staff in staff_addresses:
            bus_number = staff.get('bus_number', '')
            session = staff.get('session', '')
            halves = parse_session(session)
            
            if bus_number and bus_number.startswith('Bus'):
                bus_data[bus_number]['staff_with_addresses'] += 1
                # Staff with addresses take both AM and PM seats
                if halves['h1']:
                    bus_data[bus_number]['h1_am'] += 1
                    bus_data[bus_number]['h1_pm'] += 1
                if halves['h2']:
                    bus_data[bus_number]['h2_am'] += 1
                    bus_data[bus_number]['h2_pm'] += 1
        
        # Add capacity and staff info
        result = {}
        for bus_number in bus_data:
            data = bus_data[bus_number]
            
            # Get staff info
            if bus_number in staff_dict:
                staff = staff_dict[bus_number]
                data['capacity'] = staff.get('capacity', get_bus_capacity(bus_number))
                data['location'] = staff.get('location_name', get_bus_location(bus_number))
                data['driver'] = staff.get('driver_name', get_bus_driver(bus_number))
                data['counselor'] = staff.get('counselor_name', get_bus_counselor(bus_number))
            else:
                data['capacity'] = get_bus_capacity(bus_number)
                data['location'] = get_bus_location(bus_number)
                data['driver'] = get_bus_driver(bus_number)
                data['counselor'] = get_bus_counselor(bus_number)
            
            # Calculate available seats
            cap = data['capacity']
            data['h1_am_available'] = cap - data['h1_am']
            data['h1_pm_available'] = cap - data['h1_pm']
            data['h2_am_available'] = cap - data['h2_am']
            data['h2_pm_available'] = cap - data['h2_pm']
            
            result[bus_number] = data
        
        return {
            "status": "success",
            "buses": result
        }
    except Exception as e:
        logging.error(f"Error getting seat availability JSON: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))
@router.get("/download/seat-availability")
async def download_seat_availability():
    """Download seat availability as formatted Excel file matching the Google Sheet"""
    from fastapi.responses import Response
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    
    try:
        # Get all campers with bus assignments
        campers = await db.campers.find({
            "$or": [
                {"am_bus_number": {"$exists": True, "$nin": ["NONE", ""]}},
                {"pm_bus_number": {"$exists": True, "$nin": ["NONE", ""]}}
            ]
        }).to_list(None)
        
        # Get staff configurations from database
        staff_configs = await db.bus_staff.find({}).to_list(None)
        staff_dict = {c['bus_number']: c for c in staff_configs}
        
        # Get shadows, assigned staff, and staff with addresses for notes column
        shadows = await db.shadows.find({}).to_list(None)
        assigned_staff = await db.bus_assigned_staff.find({}).to_list(None)
        staff_addresses = await db.staff_addresses.find({}).to_list(None)
        
        # Generate cover sheet data with staff info, shadows, and assigned staff
        sheet_data = cover_sheet_generator.generate_cover_sheet(campers, staff_dict, shadows, assigned_staff, staff_addresses)
        
        # Create Excel workbook with formatting
        wb = Workbook()
        ws = wb.active
        ws.title = "Seat Availability"
        
        # Define styles
        title_font = Font(name='Arial', size=16, bold=True, color='1F4E79')
        subtitle_font = Font(name='Arial', size=12, bold=True, color='1F4E79')
        header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        data_font = Font(name='Arial', size=10)
        totals_font = Font(name='Arial', size=11, bold=True)
        totals_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        
        # Border styles
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Alternating row colors
        light_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        
        # Available column colors based on seat count
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        orange_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        green_font = Font(name='Arial', size=10, bold=True, color='006100')
        orange_font = Font(name='Arial', size=10, bold=True, color='9C5700')
        red_font = Font(name='Arial', size=10, bold=True, color='9C0006')
        
        # Available column indices (1-indexed): 7 (H1 AM Avail), 9 (H1 PM Avail), 11 (H2 AM Avail), 13 (H2 PM Avail)
        avail_cols = [7, 9, 11, 13]  # 1-indexed, removed column 14 (Available)
        
        # Write data with formatting
        for row_idx, row_data in enumerate(sheet_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Title row (row 1)
                if row_idx == 1:
                    cell.font = title_font
                    cell.alignment = Alignment(horizontal='left')
                # Subtitle row (row 2)
                elif row_idx == 2:
                    cell.font = subtitle_font
                    cell.alignment = Alignment(horizontal='left')
                # Header row (row 4)
                elif row_idx == 4:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                # TOTALS row
                elif row_data and row_data[0] == 'TOTALS':
                    cell.font = totals_font
                    cell.fill = totals_fill
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                # AVAILABLE SEATS header
                elif row_data and 'AVAILABLE SEATS' in str(row_data[0]):
                    cell.font = subtitle_font
                # Data rows (bus data)
                elif row_idx > 4 and row_data and str(row_data[0]).startswith('Bus'):
                    cell.font = data_font
                    cell.border = thin_border
                    # Alternating row colors
                    if (row_idx - 5) % 2 == 0:
                        cell.fill = light_fill
                    else:
                        cell.fill = white_fill
                    
                    # Center align numeric columns
                    if col_idx >= 5:
                        cell.alignment = Alignment(horizontal='center')
                    
                    # Color the Available columns based on value
                    if col_idx in avail_cols:
                        try:
                            avail = int(value) if value is not None else 0
                            if avail < 5:
                                cell.fill = red_fill
                                cell.font = red_font
                            elif avail <= 10:
                                cell.fill = orange_fill
                                cell.font = orange_font
                            else:
                                cell.fill = green_fill
                                cell.font = green_font
                        except:
                            pass
                # Summary rows at bottom
                elif row_data and ('Half' in str(row_data[0]) or 'Available' in str(row_data[0])):
                    cell.font = data_font
        
        # Set column widths for 13 columns
        column_widths = [10, 18, 16, 16, 7, 10, 10, 10, 10, 10, 10, 10, 10]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Freeze the header row
        ws.freeze_panes = 'A5'
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"seat_availability_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
                "Cache-Control": "no-cache, no-store, must-revalidate",
                "Pragma": "no-cache",
                "Expires": "0",
                "Access-Control-Expose-Headers": "Content-Disposition"
            }
        )
    except Exception as e:
        logging.error(f"Error generating seat availability: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))
@router.post("/update-seat-availability-sheet")
async def update_seat_availability_sheet():
    """
    Update the seat availability Google Sheet (1ZK58gjF4BO0HF_2y6oovrjzRH3qV5zAs8H-7CeKOSGE)
    with current bus assignments.
    Uses 14-column format with availability columns.
    """
    try:
        # Get all campers with bus assignments
        campers = await db.campers.find({
            "am_bus_number": {"$exists": True, "$nin": ["NONE", ""]}
        }).to_list(None)
        
        # Get staff configurations from database
        staff_configs = await db.bus_staff.find({}).to_list(None)
        staff_dict = {c['bus_number']: c for c in staff_configs}
        
        # Get shadows, assigned staff, and staff with addresses for notes column
        shadows = await db.shadows.find({}).to_list(None)
        assigned_staff = await db.bus_assigned_staff.find({}).to_list(None)
        staff_addresses = await db.staff_addresses.find({}).to_list(None)
        
        # Generate cover sheet data in 14-column format with availability columns
        sheet_data = cover_sheet_generator.generate_cover_sheet(campers, staff_dict, shadows, assigned_staff, staff_addresses)
        
        # Use dedicated seat availability webhook
        webhook_url = os.environ.get('SEAT_AVAILABILITY_WEBHOOK_URL', '')
        if not webhook_url:
            return {
                "status": "error",
                "message": "SEAT_AVAILABILITY_WEBHOOK_URL not configured"
            }
        
        payload = {
            "action": "update_seat_availability",
            "sheet_id": OUTPUT_SHEET_ID,
            "data": sheet_data
        }
        
        async with httpx.AsyncClient(timeout=60.0, follow_redirects=True) as client:
            response = await client.post(webhook_url, json=payload)
            
            if response.status_code == 200:
                return {
                    "status": "success",
                    "message": f"Updated seat availability sheet with {len(sheet_data)} rows",
                    "sheet_url": f"https://docs.google.com/spreadsheets/d/{OUTPUT_SHEET_ID}/edit"
                }
            else:
                return {
                    "status": "error",
                    "message": f"Webhook returned status {response.status_code}",
                    "response": response.text
                }
    except Exception as e:
        logging.error(f"Error updating seat availability sheet: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/push-seat-availability-to-sheet")
async def push_seat_availability_to_sheet():
    """
    Push current seat availability data to Google Sheet via webhook.
    This is the button-triggered version that shows detailed status.
    Uses 14-column format with availability columns.
    """
    try:
        # Get all campers with bus assignments
        all_campers = await db.campers.find({}).to_list(None)
        
        # Filter to campers with valid bus assignments
        campers_with_buses = [c for c in all_campers if 
            (c.get('am_bus_number', '') and c.get('am_bus_number', '') != 'NONE' and c.get('am_bus_number', '').startswith('Bus')) or
            (c.get('pm_bus_number', '') and c.get('pm_bus_number', '') != 'NONE' and c.get('pm_bus_number', '').startswith('Bus'))
        ]
        
        # Get staff configurations from database
        staff_configs = await db.bus_staff.find({}).to_list(None)
        staff_dict = {c['bus_number']: c for c in staff_configs}
        
        # Get shadows, assigned staff, and staff with addresses for notes column
        shadows = await db.shadows.find({}).to_list(None)
        assigned_staff = await db.bus_assigned_staff.find({}).to_list(None)
        staff_addresses = await db.staff_addresses.find({}).to_list(None)
        
        # Generate cover sheet data in 14-column format with availability columns and Notes
        sheet_data = cover_sheet_generator.generate_cover_sheet(campers_with_buses, staff_dict, shadows, assigned_staff, staff_addresses)
        
        # Use dedicated seat availability webhook
        webhook_url = os.environ.get('SEAT_AVAILABILITY_WEBHOOK_URL', '')
        if not webhook_url:
            return {
                "status": "error",
                "message": "SEAT_AVAILABILITY_WEBHOOK_URL not configured. Please set up the webhook."
            }
        
        payload = {
            "action": "update_seat_availability",
            "sheet_id": OUTPUT_SHEET_ID,
            "data": sheet_data
        }
        
        logging.info(f"Pushing {len(sheet_data)} rows to seat availability sheet (14 columns with Notes)")
        
        async with httpx.AsyncClient(timeout=60.0, follow_redirects=True) as client:
            response = await client.post(webhook_url, json=payload)
            response_text = response.text
            
            logging.info(f"Webhook response: {response.status_code} - {response_text}")
            
            if response.status_code == 200:
                return {
                    "status": "success",
                    "message": f"✓ Updated seat availability sheet with {len(sheet_data)} rows ({len(campers_with_buses)} campers)",
                    "sheet_url": f"https://docs.google.com/spreadsheets/d/{OUTPUT_SHEET_ID}/edit",
                    "rows_updated": len(sheet_data)
                }
            else:
                return {
                    "status": "error",
                    "message": f"Webhook returned status {response.status_code}: {response_text}",
                    "response": response_text
                }
    except Exception as e:
        logging.error(f"Error pushing seat availability: {str(e)}")
        return {
            "status": "error", 
            "message": f"Error: {str(e)}"
        }
@router.post("/update-output-sheet")
async def update_output_google_sheet():
    """
    Update the output Google Sheets document with all camper bus assignments.
    Sheet ID: 1ZK58gjF4BO0HF_2y6oovrjzRH3qV5zAs8H-7CeKOSGE
    """
    import httpx
    
    logger.info("=== UPDATING OUTPUT GOOGLE SHEET ===")
    logger.info(f"Sheet ID: {OUTPUT_SHEET_ID}")
    logger.info(f"Sheet URL: https://docs.google.com/spreadsheets/d/{OUTPUT_SHEET_ID}/edit")
    
    try:
        # Get all campers from database
        all_campers = await db.campers.find({}).to_list(None)
        logger.info(f"Found {len(all_campers)} campers in database")
        
        # Prepare data for sheet
        # Format: Name, Address, Town, Zip, Session Type, AM Bus, PM Bus
        sheet_data = []
        seen_campers = set()
        
        for camper in all_campers:
            first_name = camper.get('first_name', '')
            last_name = camper.get('last_name', '')
            camper_id = camper.get('_id', '')
            
            # Skip PM-specific entries (we'll combine data)
            if camper_id.endswith('_PM'):
                continue
            
            full_name = f"{first_name} {last_name}"
            if full_name in seen_campers:
                continue
            seen_campers.add(full_name)
            
            address = camper.get('location', {}).get('address', '')
            town = camper.get('town', '')
            zip_code = camper.get('zip_code', '')
            session = camper.get('session', camper.get('pickup_type', ''))
            am_bus = camper.get('am_bus_number', '')
            pm_bus = camper.get('pm_bus_number', '')
            
            # Clean up bus values - don't show NONE
            if am_bus == 'NONE':
                am_bus = ''
            if pm_bus == 'NONE':
                pm_bus = ''
            
            sheet_data.append({
                'name': full_name,
                'first_name': first_name,
                'last_name': last_name,
                'address': address,
                'town': town,
                'zip': zip_code,
                'session': session,
                'am_bus': am_bus,
                'pm_bus': pm_bus
            })
        
        logger.info(f"Prepared {len(sheet_data)} unique campers for sheet")
        
        # Build CSV-like data for Google Sheets
        # Headers
        headers = ['First Name', 'Last Name', 'Address', 'Town', 'Zip', 'Session Type', 'AM Bus', 'PM Bus']
        
        # Sort by last name
        sheet_data.sort(key=lambda x: (x['last_name'], x['first_name']))
        
        # Convert to rows
        rows = [headers]
        for camper in sheet_data:
            rows.append([
                camper['first_name'],
                camper['last_name'],
                camper['address'],
                camper['town'],
                camper['zip'],
                camper['session'],
                camper['am_bus'],
                camper['pm_bus']
            ])
        
        # Use the webhook URL to update the sheet
        webhook_url = os.environ.get('GOOGLE_SHEETS_WEBHOOK_URL', '')
        
        if webhook_url:
            # Try to use webhook for update
            logger.info("Attempting update via webhook...")
            
            async with httpx.AsyncClient(timeout=60.0) as client:
                try:
                    response = await client.post(
                        webhook_url,
                        json={
                            'action': 'updateOutputSheet',
                            'sheetId': OUTPUT_SHEET_ID,
                            'data': rows
                        }
                    )
                    
                    if response.status_code == 200:
                        logger.info("✓ Sheet updated via webhook")
                        return {
                            "status": "success",
                            "message": f"Updated {len(sheet_data)} campers in Google Sheet",
                            "sheet_url": f"https://docs.google.com/spreadsheets/d/{OUTPUT_SHEET_ID}/edit",
                            "rows_written": len(rows),
                            "method": "webhook"
                        }
                except Exception as e:
                    logger.warning(f"Webhook update failed: {str(e)}, trying direct API...")
        
        # If no webhook or webhook failed, try direct API access
        # This requires a service account with access to the sheet
        try:
            from google.oauth2 import service_account
            from googleapiclient.discovery import build
            
            # Check for service account credentials
            creds_path = os.environ.get('GOOGLE_APPLICATION_CREDENTIALS', '/app/backend/service-account.json')
            
            if os.path.exists(creds_path):
                credentials = service_account.Credentials.from_service_account_file(
                    creds_path,
                    scopes=['https://www.googleapis.com/auth/spreadsheets']
                )
                
                service = build('sheets', 'v4', credentials=credentials)
                
                # Clear existing data
                logger.info("Clearing existing data...")
                service.spreadsheets().values().clear(
                    spreadsheetId=OUTPUT_SHEET_ID,
                    range='Sheet1!A1:H1000'
                ).execute()
                
                # Write new data
                logger.info(f"Writing {len(rows)} rows...")
                result = service.spreadsheets().values().update(
                    spreadsheetId=OUTPUT_SHEET_ID,
                    range='Sheet1!A1',
                    valueInputOption='USER_ENTERED',
                    body={'values': rows}
                ).execute()
                
                logger.info(f"✓ Updated {result.get('updatedCells', 0)} cells")
                
                return {
                    "status": "success",
                    "message": f"Updated {len(sheet_data)} campers in Google Sheet",
                    "sheet_url": f"https://docs.google.com/spreadsheets/d/{OUTPUT_SHEET_ID}/edit",
                    "rows_written": len(rows),
                    "cells_updated": result.get('updatedCells', 0),
                    "method": "direct_api"
                }
            else:
                logger.warning("No service account credentials found")
                
        except ImportError:
            logger.warning("Google API client not fully configured")
        except Exception as e:
            logger.error(f"Direct API update failed: {str(e)}")
        
        # Return data for manual update if automated methods fail
        return {
            "status": "manual_required",
            "message": "Automated update not available. Use the data below to update manually.",
            "sheet_url": f"https://docs.google.com/spreadsheets/d/{OUTPUT_SHEET_ID}/edit",
            "total_campers": len(sheet_data),
            "headers": headers,
            "sample_data": rows[:10],
            "full_data_available": True,
            "instructions": "Copy the data from /api/export-campers-csv endpoint to update the sheet manually"
        }
        
    except Exception as e:
        logger.error(f"Error updating output sheet: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/export-campers-csv")
async def export_campers_csv():
    """Export all campers as CSV for manual sheet update"""
    from fastapi.responses import StreamingResponse
    import io
    import csv
    
    # Get all campers
    all_campers = await db.campers.find({}).to_list(None)
    
    # Prepare CSV
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Headers
    writer.writerow(['First Name', 'Last Name', 'Address', 'Town', 'Zip', 'Session Type', 'AM Bus', 'PM Bus'])
    
    seen = set()
    for camper in sorted(all_campers, key=lambda x: (x.get('last_name', ''), x.get('first_name', ''))):
        camper_id = camper.get('_id', '')
        if camper_id.endswith('_PM'):
            continue
        
        name = f"{camper.get('first_name', '')} {camper.get('last_name', '')}"
        if name in seen:
            continue
        seen.add(name)
        
        am_bus = camper.get('am_bus_number', '')
        pm_bus = camper.get('pm_bus_number', '')
        if am_bus == 'NONE':
            am_bus = ''
        if pm_bus == 'NONE':
            pm_bus = ''
        
        writer.writerow([
            camper.get('first_name', ''),
            camper.get('last_name', ''),
            camper.get('location', {}).get('address', ''),
            camper.get('town', ''),
            camper.get('zip_code', ''),
            camper.get('session', camper.get('pickup_type', '')),
            am_bus,
            pm_bus
        ])
    
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=camper_bus_assignments.csv"}
    )
@router.get("/download/bus-assignments")
async def download_bus_assignments():
    """Download bus assignments as CSV with AM and PM bus columns"""
    from fastapi.responses import Response
    from io import StringIO
    import csv as csv_module
    
    try:
        campers = await db.campers.find({
            "$or": [
                {"am_bus_number": {"$exists": True, "$nin": ["NONE", ""]}},
                {"pm_bus_number": {"$exists": True, "$nin": ["NONE", ""]}}
            ]
        }).to_list(None)
        
        # Create CSV
        output = StringIO()
        writer = csv_module.writer(output)
        
        # Simple header - just name and bus numbers
        writer.writerow([
            'Last Name',
            'First Name', 
            'AM Bus',
            'PM Bus'
        ])
        
        # Track campers to avoid duplicates
        seen_campers = set()
        
        # Data rows - sorted by last name, first name
        for camper in sorted(campers, key=lambda x: (x.get('last_name', '').lower(), x.get('first_name', '').lower())):
            camper_id = camper.get('_id', '')
            
            # Skip _PM suffix entries
            if str(camper_id).endswith('_PM'):
                continue
            
            camper_key = f"{camper.get('last_name', '')}_{camper.get('first_name', '')}"
            if camper_key in seen_campers:
                continue
            seen_campers.add(camper_key)
            
            am_bus = camper.get('am_bus_number', '')
            pm_bus = camper.get('pm_bus_number', '')
            
            # Look for _PM entry for this camper (may have different PM bus)
            for c in campers:
                if str(c.get('_id', '')).endswith('_PM'):
                    if c.get('first_name') == camper.get('first_name') and c.get('last_name') == camper.get('last_name'):
                        pm_bus = c.get('pm_bus_number', pm_bus)
                        break
            
            # Display "NONE" as "N/A"
            if am_bus == 'NONE' or not am_bus:
                am_bus = 'N/A'
            if pm_bus == 'NONE' or not pm_bus:
                pm_bus = 'N/A'
            
            writer.writerow([
                camper.get('last_name', ''),
                camper.get('first_name', ''),
                am_bus,
                pm_bus
            ])
        
        output.seek(0)
        
        filename = f"bus_assignments_{datetime.now().strftime('%Y%m%d')}.csv"
        
        return Response(
            content=output.getvalue(),
            media_type="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
                "Content-Type": "text/csv; charset=utf-8",
                "Cache-Control": "no-cache, no-store, must-revalidate",
                "Pragma": "no-cache",
                "Expires": "0",
                "Access-Control-Expose-Headers": "Content-Disposition"
            }
        )
    except Exception as e:
        logging.error(f"Error generating download: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


