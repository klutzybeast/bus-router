from fastapi import FastAPI, APIRouter, HTTPException
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone
import requests
import httpx
import googlemaps
from io import StringIO
import csv
from route_optimizer import RouteOptimizer
from campminder_integration import CampMinderAPI
from sheets_generator import SheetsDataGenerator
from cover_sheet_generator import CoverSheetGenerator
from route_printer import RoutePrinter
from sibling_offset import apply_sibling_offset
from bus_config import get_bus_info, get_all_buses, get_camp_address, get_bus_home_location, get_bus_capacity, get_bus_driver, get_bus_counselor, get_bus_location
from contextlib import asynccontextmanager
import asyncio

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection with Atlas-compatible settings
mongo_url = os.environ.get('MONGO_URL')
db_name = os.environ.get('DB_NAME')

if not mongo_url:
    logging.error("MONGO_URL environment variable not set!")
    raise ValueError("MONGO_URL environment variable is required")

if not db_name:
    logging.error("DB_NAME environment variable not set!")
    raise ValueError("DB_NAME environment variable is required")

# Check if this is an Atlas connection (contains mongodb+srv or mongodb.net)
is_atlas = 'mongodb.net' in mongo_url or 'mongodb+srv' in mongo_url

# Log connection type
logging.info(f"MongoDB connection type: {'Atlas' if is_atlas else 'Local'}")
logging.info(f"Database name: {db_name}")

try:
    if is_atlas:
        # Atlas connection - use SRV-compatible settings
        client = AsyncIOMotorClient(
            mongo_url,
            serverSelectionTimeoutMS=120000,
            connectTimeoutMS=60000,
            socketTimeoutMS=120000,
            retryWrites=True,
            retryReads=True,
            maxPoolSize=50,
            minPoolSize=0,
            maxIdleTimeMS=60000,
            waitQueueTimeoutMS=120000,
            appName="BusRoutingApp",
            directConnection=False,
        )
    else:
        # Local MongoDB connection
        client = AsyncIOMotorClient(
            mongo_url,
            serverSelectionTimeoutMS=30000,
            connectTimeoutMS=20000,
            socketTimeoutMS=30000,
            retryWrites=True,
            retryReads=True,
            maxPoolSize=10,
            minPoolSize=1
        )
    
    db = client[db_name]
    db_connected = False  # Will be set to True when first successful connection
    logging.info("MongoDB client initialized successfully")
except Exception as e:
    logging.error(f"Failed to initialize MongoDB client: {str(e)}")
    client = None
    db = None
    db_connected = False

# Initialize services with error handling
try:
    google_maps_key = os.environ.get('GOOGLE_MAPS_API_KEY', '')
    if google_maps_key:
        gmaps = googlemaps.Client(key=google_maps_key)
        logging.info("Google Maps client initialized")
    else:
        gmaps = None
        logging.warning("GOOGLE_MAPS_API_KEY not set - geocoding will be disabled")
except Exception as e:
    gmaps = None
    logging.error(f"Failed to initialize Google Maps client: {e}")

# PositionStack API key (free backup geocoding)
POSITIONSTACK_API_KEY = os.environ.get('POSITIONSTACK_API_KEY', '')
if POSITIONSTACK_API_KEY:
    logging.info("PositionStack API key configured as backup geocoder")
else:
    logging.warning("POSITIONSTACK_API_KEY not set - no backup geocoding available")

route_optimizer = RouteOptimizer(num_buses=34)
sheets_generator = SheetsDataGenerator()
cover_sheet_generator = CoverSheetGenerator()
route_printer = RoutePrinter(gmaps) if gmaps else None
campminder_api = CampMinderAPI(
    api_key=os.environ.get('CAMPMINDER_API_KEY', ''),
    subscription_key=os.environ.get('CAMPMINDER_SUBSCRIPTION_KEY', '')
)

app = FastAPI()
api_router = APIRouter(prefix="/api")

# Auto-sync configuration
AUTO_SYNC_ENABLED = os.environ.get('AUTO_SYNC_ENABLED', 'true').lower() == 'true'
SYNC_INTERVAL_MINUTES = int(os.environ.get('SYNC_INTERVAL_MINUTES', '15'))
CAMPMINDER_SHEET_ID = os.environ.get('CAMPMINDER_SHEET_ID', '')
# Output sheet for bus assignments (different from source sheet)
OUTPUT_SHEET_ID = os.environ.get('OUTPUT_SHEET_ID', '1ZK58gjF4BO0HF_2y6oovrjzRH3qV5zAs8H-7CeKOSGE')
sync_task = None
last_sync_time = None

BUS_COLORS = [
    "#FF0000",   # Bus 1 - Pure Red
    "#228B22",   # Bus 2 - Forest Green (darker green)
    "#0000FF",   # Bus 3 - Pure Blue
    "#B8860B",   # Bus 4 - Dark Goldenrod (darker yellow)
    "#FF00FF",   # Bus 5 - Pure Magenta
    "#008B8B",   # Bus 6 - Dark Cyan (darker)
    "#800000",   # Bus 7 - Dark Red (Maroon)
    "#008000",   # Bus 8 - Dark Green
    "#000080",   # Bus 9 - Dark Blue (Navy)
    "#808000",   # Bus 10 - Olive
    "#800080",   # Bus 11 - Purple
    "#008080",   # Bus 12 - Teal
    "#FFA500",   # Bus 13 - Orange
    "#FF1493",   # Bus 14 - Deep Pink
    "#00CED1",   # Bus 15 - Dark Turquoise
    "#FF4500",   # Bus 16 - Orange Red
    "#9400D3",   # Bus 17 - Dark Violet
    "#32CD32",   # Bus 18 - Lime Green
    "#DC143C",   # Bus 19 - Crimson
    "#4169E1",   # Bus 20 - Royal Blue
    "#FF8C00",   # Bus 21 - Dark Orange
    "#8B4513",   # Bus 22 - Saddle Brown
    "#6B8E23",   # Bus 23 - Olive Drab (darker chartreuse)
    "#2E8B57",   # Bus 24 - Sea Green (darker spring green)
    "#FF69B4",   # Bus 25 - Hot Pink
    "#4682B4",   # Bus 26 - Steel Blue
    "#D2691E",   # Bus 27 - Chocolate
    "#DAA520",   # Bus 28 - Goldenrod (darker gold)
    "#8A2BE2",   # Bus 29 - Blue Violet
    "#5F9EA0",   # Bus 30 - Cadet Blue
    "#A52A2A",   # Bus 31 - Brown
    "#DEB887",   # Bus 32 - Burlywood
    "#6495ED",   # Bus 33 - Cornflower Blue
    "#FF7F50"    # Bus 34 - Coral
]

class GeoLocation(BaseModel):
    latitude: float
    longitude: float
    address: Optional[str] = None

class CamperPin(BaseModel):
    first_name: str
    last_name: str
    location: GeoLocation
    am_bus_number: str
    pm_bus_number: str
    bus_color: str
    session: str
    pickup_type: str
    town: Optional[str] = None
    zip_code: Optional[str] = None

def get_bus_color(bus_number: str) -> str:
    try:
        bus_num = int(''.join(filter(str.isdigit, bus_number)))
        return BUS_COLORS[(bus_num - 1) % len(BUS_COLORS)]
    except (ValueError, IndexError):
        return BUS_COLORS[0]

# In-memory cache for current session (reduces DB lookups)
_geocode_memory_cache = {}

def normalize_address(address: str, town: str = "", zip_code: str = "") -> str:
    """Normalize address for consistent cache keys"""
    full = f"{address}, {town}, {zip_code}" if town else address
    # Normalize: lowercase, strip extra spaces, remove common variations
    normalized = ' '.join(full.lower().split())
    return normalized

async def get_cached_geocode(address_key: str) -> Optional[dict]:
    """Check MongoDB cache for geocoded address"""
    try:
        # Check memory cache first
        if address_key in _geocode_memory_cache:
            return _geocode_memory_cache[address_key]
        
        # Check database cache
        cached = await db.geocode_cache.find_one({"address_key": address_key})
        if cached and cached.get('latitude') and cached.get('longitude'):
            # Store in memory cache for faster subsequent lookups
            _geocode_memory_cache[address_key] = {
                'latitude': cached['latitude'],
                'longitude': cached['longitude'],
                'formatted_address': cached.get('formatted_address', '')
            }
            return _geocode_memory_cache[address_key]
        return None
    except Exception as e:
        logging.error(f"Cache lookup error: {e}")
        return None

async def save_geocode_cache(address_key: str, latitude: float, longitude: float, formatted_address: str, source: str):
    """Save geocoded result to MongoDB cache"""
    try:
        await db.geocode_cache.update_one(
            {"address_key": address_key},
            {
                "$set": {
                    "address_key": address_key,
                    "latitude": latitude,
                    "longitude": longitude,
                    "formatted_address": formatted_address,
                    "source": source,
                    "cached_at": datetime.now(timezone.utc)
                }
            },
            upsert=True
        )
        # Also update memory cache
        _geocode_memory_cache[address_key] = {
            'latitude': latitude,
            'longitude': longitude,
            'formatted_address': formatted_address
        }
        logging.info(f"Cached geocode for: {address_key[:50]}... (source: {source})")
    except Exception as e:
        logging.error(f"Cache save error: {e}")

def geocode_with_google(full_address: str) -> Optional[dict]:
    """Geocode using Google Maps API"""
    if not gmaps:
        return None
    try:
        result = gmaps.geocode(full_address)
        if result and len(result) > 0:
            location = result[0]['geometry']['location']
            return {
                'latitude': location['lat'],
                'longitude': location['lng'],
                'formatted_address': result[0]['formatted_address']
            }
    except Exception as e:
        logging.error(f"Google geocoding error for {full_address}: {str(e)}")
    return None

def geocode_with_positionstack(full_address: str) -> Optional[dict]:
    """Geocode using PositionStack API (free backup)"""
    if not POSITIONSTACK_API_KEY:
        return None
    try:
        response = requests.get(
            "http://api.positionstack.com/v1/forward",
            params={
                "access_key": POSITIONSTACK_API_KEY,
                "query": full_address,
                "limit": 1
            },
            timeout=10
        )
        if response.status_code == 200:
            data = response.json()
            if data.get('data') and len(data['data']) > 0:
                result = data['data'][0]
                return {
                    'latitude': result['latitude'],
                    'longitude': result['longitude'],
                    'formatted_address': result.get('label', full_address)
                }
    except Exception as e:
        logging.error(f"PositionStack geocoding error for {full_address}: {str(e)}")
    return None

def geocode_address(address: str, town: str = "", zip_code: str = "") -> Optional[GeoLocation]:
    """
    Synchronous geocode function (for backward compatibility).
    Note: This doesn't use cache - use geocode_address_cached for cached version.
    """
    full_address = f"{address}, {town}, {zip_code}" if town else address
    if not full_address.strip():
        return None
    
    # Try Google first
    result = geocode_with_google(full_address)
    if result:
        return GeoLocation(
            latitude=result['latitude'],
            longitude=result['longitude'],
            address=result['formatted_address']
        )
    
    # Try PositionStack as backup
    result = geocode_with_positionstack(full_address)
    if result:
        return GeoLocation(
            latitude=result['latitude'],
            longitude=result['longitude'],
            address=result['formatted_address']
        )
    
    return None

async def geocode_address_cached(address: str, town: str = "", zip_code: str = "") -> Optional[GeoLocation]:
    """
    Geocode address with caching - ALWAYS use this for batch operations.
    1. Check cache first
    2. If not cached, try Google
    3. If Google fails, try PositionStack
    4. Save result to cache
    """
    full_address = f"{address}, {town}, {zip_code}" if town else address
    if not full_address.strip():
        return None
    
    address_key = normalize_address(address, town, zip_code)
    
    # Step 1: Check cache
    cached = await get_cached_geocode(address_key)
    if cached:
        logging.debug(f"Cache HIT for: {address_key[:50]}...")
        return GeoLocation(
            latitude=cached['latitude'],
            longitude=cached['longitude'],
            address=cached.get('formatted_address', full_address)
        )
    
    logging.info(f"Cache MISS - geocoding: {address_key[:50]}...")
    
    # Step 2: Try Google
    result = geocode_with_google(full_address)
    if result:
        await save_geocode_cache(address_key, result['latitude'], result['longitude'], result['formatted_address'], 'google')
        return GeoLocation(
            latitude=result['latitude'],
            longitude=result['longitude'],
            address=result['formatted_address']
        )
    
    # Step 3: Try PositionStack as backup
    result = geocode_with_positionstack(full_address)
    if result:
        await save_geocode_cache(address_key, result['latitude'], result['longitude'], result['formatted_address'], 'positionstack')
        return GeoLocation(
            latitude=result['latitude'],
            longitude=result['longitude'],
            address=result['formatted_address']
        )
    
    logging.warning(f"All geocoding failed for: {full_address}")
    return None

@api_router.get("/")
async def root():
    return {"message": "Bus Routing API"}

# Health check endpoint on API router
@api_router.get("/health")
async def api_health_check():
    """Health check endpoint for API"""
    return {"status": "healthy", "service": "bus-routing-api"}

# Geocode cache stats endpoint
@api_router.get("/geocode-cache-stats")
async def get_geocode_cache_stats():
    """Get statistics about the geocoding cache"""
    try:
        total_cached = await db.geocode_cache.count_documents({})
        google_cached = await db.geocode_cache.count_documents({"source": "google"})
        positionstack_cached = await db.geocode_cache.count_documents({"source": "positionstack"})
        memory_cache_size = len(_geocode_memory_cache)
        
        return {
            "status": "success",
            "total_cached_addresses": total_cached,
            "by_source": {
                "google": google_cached,
                "positionstack": positionstack_cached
            },
            "memory_cache_size": memory_cache_size,
            "message": f"Cache has {total_cached} addresses. New addresses will use Google first, PositionStack as backup."
        }
    except Exception as e:
        return {"status": "error", "error": str(e)}

# Config check endpoint (for debugging deployment issues)
@api_router.get("/config-check")
async def config_check():
    """Check if critical environment variables are configured"""
    webhook_url = os.environ.get('GOOGLE_SHEETS_WEBHOOK_URL', '')
    return {
        "webhook_configured": bool(webhook_url),
        "webhook_url_preview": webhook_url[:50] + "..." if len(webhook_url) > 50 else webhook_url if webhook_url else "NOT SET",
        "positionstack_configured": bool(os.environ.get('POSITIONSTACK_API_KEY', '')),
        "google_maps_configured": bool(os.environ.get('GOOGLE_MAPS_API_KEY', ''))
    }

# Database status endpoint on API router
@api_router.get("/db-status")
async def api_db_status():
    """Check database connection status"""
    global db_connected
    if db is None:
        return {"status": "error", "error": "Database not configured"}
    try:
        await asyncio.wait_for(db.command('ping'), timeout=10.0)
        camper_count = await asyncio.wait_for(db.campers.count_documents({}), timeout=10.0)
        db_connected = True
        return {
            "status": "connected",
            "camper_count": camper_count,
            "db_type": "atlas" if is_atlas else "local"
        }
    except asyncio.TimeoutError:
        db_connected = False
        return {"status": "timeout", "error": "Database connection timed out"}
    except Exception as e:
        db_connected = False
        return {"status": "error", "error": str(e)}

# Force sync endpoint on API router
@api_router.post("/force-sync")
async def api_force_sync():
    """Force a sync from Google Sheets"""
    if db is None:
        return {"status": "error", "error": "Database not configured"}
    try:
        await asyncio.wait_for(db.command('ping'), timeout=10.0)
        await auto_sync_campminder()
        camper_count = await db.campers.count_documents({})
        return {"status": "success", "camper_count": camper_count}
    except Exception as e:
        return {"status": "error", "error": str(e)}

@api_router.get("/campers")
async def get_campers():
    global db_connected
    try:
        # Return campers with valid locations and at least one valid bus assignment
        existing_campers = await asyncio.wait_for(
            db.campers.find({
                "location.latitude": {"$ne": 0.0},
                "$or": [
                    {"am_bus_number": {"$regex": "^Bus"}},
                    {"pm_bus_number": {"$regex": "^Bus"}}
                ]
            }).to_list(None),
            timeout=30.0  # 30 second timeout
        )
        
        db_connected = True
        
        # Convert _id to string and clean up data
        result = []
        for camper in existing_campers:
            am_bus = camper.get('am_bus_number', '')
            pm_bus = camper.get('pm_bus_number', '')
            
            # Clean up NONE values - replace with empty string
            if am_bus == 'NONE' or not am_bus.startswith('Bus'):
                am_bus = ''
            if pm_bus == 'NONE' or not pm_bus.startswith('Bus'):
                pm_bus = ''
            
            # Skip campers with no valid bus at all
            if not am_bus and not pm_bus:
                continue
            
            camper_dict = {
                "_id": str(camper['_id']),
                "first_name": camper.get('first_name', ''),
                "last_name": camper.get('last_name', ''),
                "location": camper.get('location', {}),
                "am_bus_number": am_bus,
                "pm_bus_number": pm_bus,
                "bus_number": am_bus or pm_bus,  # For compatibility - use whichever is valid
                "bus_color": camper.get('bus_color', ''),
                "session": camper.get('session', ''),
                "pickup_type": camper.get('pickup_type', ''),
                "town": camper.get('town', ''),
                "zip_code": camper.get('zip_code', '')
            }
            result.append(camper_dict)
        
        return result
    except asyncio.TimeoutError:
        logging.error("Timeout fetching campers - database may be slow or unavailable")
        raise HTTPException(status_code=503, detail="Database timeout - please try again")
    except Exception as e:
        logging.error(f"Error fetching campers: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/campers/needs-address")
async def get_campers_needing_address():
    """Get campers who have bus assignments but no address"""
    try:
        campers = await db.campers.find({
            "location.latitude": 0.0,
            "$or": [
                {"am_bus_number": {"$exists": True, "$regex": "^Bus"}},
                {"pm_bus_number": {"$exists": True, "$regex": "^Bus"}}
            ]
        }).to_list(None)
        
        result = []
        for camper in campers:
            result.append({
                "_id": str(camper['_id']),
                "first_name": camper.get('first_name', ''),
                "last_name": camper.get('last_name', ''),
                "am_bus_number": camper.get('am_bus_number', ''),
                "pm_bus_number": camper.get('pm_bus_number', ''),
                "pickup_type": camper.get('pickup_type', '')
            })
        
        return result
    except Exception as e:
        logging.error(f"Error fetching campers needing address: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


class ManualCamperInput(BaseModel):
    first_name: str
    last_name: str
    address: str
    town: str
    zip_code: str
    am_bus_number: Optional[str] = "NONE"
    pm_bus_number: Optional[str] = None
    session: Optional[str] = ""

@api_router.post("/campers/add")
async def add_camper_manually(camper: ManualCamperInput):
    """Manually add a camper to the map"""
    try:
        # Geocode the address (using cached version)
        location = await geocode_address_cached(camper.address, camper.town, camper.zip_code)
        if not location:
            raise HTTPException(status_code=400, detail=f"Could not geocode address: {camper.address}, {camper.town}, {camper.zip_code}")
        
        # Generate camper ID
        camper_id = f"{camper.last_name}_{camper.first_name}_{camper.zip_code}".replace(' ', '_')
        
        # Check for existing campers at same location for offset
        existing_at_address = await db.campers.count_documents({
            "location.latitude": {"$gte": location.latitude - 0.001, "$lte": location.latitude + 0.001},
            "location.longitude": {"$gte": location.longitude - 0.001, "$lte": location.longitude + 0.001}
        })
        offset = existing_at_address * 0.00002
        
        # Determine bus values
        am_bus = camper.am_bus_number if camper.am_bus_number else "NONE"
        pm_bus = camper.pm_bus_number if camper.pm_bus_number else am_bus
        
        # Determine color and pickup type
        if am_bus == "NONE":
            bus_color = "#808080"
            pickup_type = "NEEDS BUS"
        else:
            bus_color = get_bus_color(am_bus)
            pickup_type = "AM & PM"
        
        camper_doc = {
            "_id": camper_id,
            "first_name": camper.first_name,
            "last_name": camper.last_name,
            "session": camper.session or "",
            "location": {
                "latitude": location.latitude + offset,
                "longitude": location.longitude + offset,
                "address": location.address
            },
            "town": camper.town,
            "zip_code": camper.zip_code,
            "pickup_type": pickup_type,
            "am_bus_number": am_bus,
            "pm_bus_number": pm_bus,
            "bus_color": bus_color,
            "created_at": datetime.now(timezone.utc),
            "manually_added": True
        }
        
        result = await db.campers.replace_one({"_id": camper_id}, camper_doc, upsert=True)
        
        # Trigger instant update to Google Sheet for this camper
        webhook_url = os.environ.get('GOOGLE_SHEETS_WEBHOOK_URL', '')
        if webhook_url and am_bus != "NONE":
            try:
                payload = {
                    "action": "update_camper",
                    "first_name": camper.first_name,
                    "last_name": camper.last_name,
                    "am_bus": am_bus,
                    "pm_bus": pm_bus
                }
                async with httpx.AsyncClient(timeout=30.0) as client:
                    await client.post(webhook_url, json=payload)
                    logger.info(f"✓ Google Sheet updated instantly for {camper.first_name} {camper.last_name}")
            except Exception as e:
                logger.warning(f"Failed to update sheet instantly: {str(e)}")
        
        return {
            "success": True,
            "message": f"Added {camper.first_name} {camper.last_name}",
            "camper_id": camper_id,
            "location": {"latitude": location.latitude + offset, "longitude": location.longitude + offset},
            "was_update": result.modified_count > 0
        }
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error adding camper: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.delete("/campers/{camper_id}")
async def delete_camper(camper_id: str):
    """Delete a camper from the database"""
    try:
        # URL decode the camper_id
        import urllib.parse
        decoded_id = urllib.parse.unquote(camper_id)
        
        result = await db.campers.delete_one({"_id": decoded_id})
        
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Camper not found")
        
        # Also delete any PM-specific entry
        await db.campers.delete_many({"_id": {"$regex": f"^{decoded_id}_PM"}})
        
        return {"success": True, "message": "Camper deleted"}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error deleting camper: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))



@api_router.post("/sync-campers")
async def sync_campers(csv_data: Dict[str, Any]):
    try:
        pins = []
        csv_content = csv_data.get('csv_content', '')
        
        if not csv_content:
            raise HTTPException(status_code=400, detail="No CSV content provided")
        
        csv_file = StringIO(csv_content)
        # Remove BOM if present
        if csv_content.startswith('\ufeff'):
            csv_content = csv_content[1:]
            csv_file = StringIO(csv_content)
        
        reader = csv.DictReader(csv_file)
        
        for row in reader:
            am_method = row.get('Trans-AMDropOffMethod', '')
            
            if 'AM Bus' not in am_method:
                continue
            
            am_bus = row.get('2026Transportation M AM Bus', '').strip()
            pm_bus = row.get('2026Transportation M PM Bus', '').strip()
            
            # If no bus assigned, set to NONE (but still import the camper)
            if not am_bus or 'NONE' in am_bus.upper():
                am_bus = 'NONE'
            
            first_name = row.get('First Name', '')
            last_name = row.get('Last Name', '')
            session = row.get('Enrolled Child Sessions', '')
            
            am_address = row.get('Trans-PickUpAddress', '')
            am_town = row.get('Trans-PickUpTown', '')
            am_zip = row.get('Trans-PickUpZip', '')
            
            pm_address = row.get('Trans-DropOffAddress', '')
            pm_town = row.get('Trans-DropOffTown', '')
            pm_zip = row.get('Trans-DropOffZip', '')
            
            # Determine final PM values
            final_pm_bus = pm_bus.strip() if pm_bus and pm_bus.strip() else am_bus
            if final_pm_bus and any(x in final_pm_bus.upper() for x in ['MAIN TENT', 'HOCKEY RINK', 'AUDITORIUM', 'NONE']):
                final_pm_bus = am_bus
            
            pm_final_address = pm_address if pm_address.strip() else am_address
            pm_final_town = pm_town if pm_town.strip() else am_town
            pm_final_zip = pm_zip if pm_zip.strip() else am_zip
            
            # Process ALL campers (including those with NONE bus)
            if am_address.strip():
                location = await geocode_address_cached(am_address, am_town, am_zip)
                if not location:
                    location = GeoLocation(latitude=0.0, longitude=0.0, address=f"GEOCODING FAILED: {am_address}")
                    logger.warning(f"Geocoding failed for {first_name} {last_name}: {am_address}")
                
                # Count existing in CURRENT batch (pins list) for offset
                existing_count = len([p for p in pins if 
                    abs(p.location.latitude - location.latitude) < 0.0001 and
                    abs(p.location.longitude - location.longitude) < 0.0001
                ])
                offset = existing_count * 0.00002  # Reduced to ~6 feet per sibling
                
                # Set bus color - gray for NONE, otherwise use bus color
                bus_color = "#808080" if am_bus == "NONE" else get_bus_color(am_bus)
                
                pins.append(CamperPin(
                    first_name=first_name,
                    last_name=last_name,
                    location=GeoLocation(
                        latitude=location.latitude + offset,
                        longitude=location.longitude + offset,
                        address=location.address
                    ),
                    am_bus_number=am_bus,
                    pm_bus_number=final_pm_bus,
                    bus_color=bus_color,
                    session=session,
                    pickup_type="AM & PM" if am_bus != "NONE" else "NEEDS BUS",
                    town=am_town,
                    zip_code=am_zip
                ))
            else:
                # No address - still add for tracking
                bus_color = "#808080" if am_bus == "NONE" else get_bus_color(am_bus)
                pins.append(CamperPin(
                    first_name=first_name,
                    last_name=last_name,
                    location=GeoLocation(latitude=0.0, longitude=0.0, address="ADDRESS NEEDED"),
                    am_bus_number=am_bus,
                    pm_bus_number=final_pm_bus,
                    bus_color=bus_color,
                    session=session,
                    pickup_type="NO ADDRESS",
                    town=am_town or "UNKNOWN",
                    zip_code=am_zip or "UNKNOWN"
                ))
            
            # Add separate PM pin only if has assigned bus and address is different
            if am_bus != "NONE" and pm_final_address != am_address and pm_final_address.strip():
                    location_pm = await geocode_address_cached(pm_final_address, pm_final_town, pm_final_zip)
                    if not location_pm:
                        # Geocoding failed - use placeholder
                        location_pm = GeoLocation(latitude=0.0, longitude=0.0, address=f"GEOCODING FAILED: {pm_final_address}")
                        logger.warning(f"PM geocoding failed for {first_name} {last_name}: {pm_final_address}")
                    
                    # Count existing at PM location
                    existing_pm_count = len([p for p in pins if 
                        abs(p.location.latitude - location_pm.latitude) < 0.001 and
                        abs(p.location.longitude - location_pm.longitude) < 0.001
                    ])
                    pm_offset = existing_pm_count * 0.00008
                    
                    pins.append(CamperPin(
                        first_name=first_name,
                        last_name=last_name,
                        location=GeoLocation(
                            latitude=location_pm.latitude + pm_offset,
                            longitude=location_pm.longitude + pm_offset,
                            address=location_pm.address
                        ),
                        am_bus_number=am_bus,
                        pm_bus_number=final_pm_bus,
                        bus_color=get_bus_color(final_pm_bus),
                        session=session,
                        pickup_type="PM Drop-off Only",
                        town=pm_final_town,
                        zip_code=pm_final_zip
                    ))
        
        await db.campers.delete_many({})
        
        if pins:
            pins_dict = [pin.model_dump() for pin in pins]
            await db.campers.insert_many(pins_dict)
        
        # AUTOMATICALLY apply sibling offset after inserting
        await apply_sibling_offset(db)
        
        return {"status": "success", "count": len(pins)}
    except Exception as e:
        logging.error(f"Error syncing campers: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/refresh-colors")
async def refresh_colors():
    try:
        campers = await db.campers.find({}).to_list(None)
        
        for camper in campers:
            # Use am_bus_number as primary, fallback to bus_number for legacy
            bus_num = camper.get('am_bus_number') or camper.get('bus_number', '')
            if bus_num and bus_num != 'NONE' and bus_num.startswith('Bus'):
                new_color = get_bus_color(bus_num)
            else:
                # Use PM bus if AM is not valid
                pm_bus = camper.get('pm_bus_number', '')
                if pm_bus and pm_bus != 'NONE' and pm_bus.startswith('Bus'):
                    new_color = get_bus_color(pm_bus)
                else:
                    new_color = "#808080"  # Gray for no bus
            
            await db.campers.update_one(
                {"_id": camper["_id"]},
                {"$set": {"bus_color": new_color}}
            )
        
        return {"status": "success", "updated": len(campers)}
    except Exception as e:
        logging.error(f"Error refreshing colors: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/optimize-routes")
async def optimize_routes():
    """Automatically optimize bus routes for all campers"""
    try:
        # Get all campers without bus assignments
        campers = await db.campers.find({}).to_list(None)
        
        if not campers:
            return {"status": "success", "message": "No campers to optimize"}
        
        # Run route optimization
        optimized_routes = route_optimizer.optimize_routes(campers)
        
        # Rebalance for efficiency
        balanced_routes = route_optimizer.rebalance_routes(optimized_routes)
        
        # Update database with assignments
        updates = []
        for bus_num, route in balanced_routes.items():
            for camper_data in route:
                camper_id = camper_data['camper_id']
                bus_number_str = f"Bus #{bus_num:02d}"
                
                await db.campers.update_one(
                    {"_id": camper_id},
                    {"$set": {
                        "bus_number": bus_number_str,
                        "bus_color": get_bus_color(bus_number_str)
                    }}
                )
                
                updates.append({
                    "camper_id": camper_id,
                    "bus_number": bus_num
                })
        
        return {
            "status": "success",
            "optimized_buses": len(balanced_routes),
            "assigned_campers": len(updates),
            "routes": {f"Bus #{k:02d}": len(v) for k, v in balanced_routes.items()}
        }
    except Exception as e:
        logging.error(f"Error optimizing routes: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/auto-assign-new-camper")
async def auto_assign_new_camper(camper_id: str):
    """Automatically assign optimal bus to a new camper"""
    try:
        # Get the new camper
        new_camper = await db.campers.find_one({"_id": camper_id})
        if not new_camper:
            raise HTTPException(status_code=404, detail="Camper not found")
        
        # Get existing bus routes
        all_campers = await db.campers.find({"bus_number": {"$exists": True}}).to_list(None)
        existing_routes = {}
        
        for camper in all_campers:
            bus_num_str = camper.get('bus_number', '')
            if bus_num_str:
                # Extract bus number
                bus_num = int(''.join(filter(str.isdigit, bus_num_str)))
                if bus_num not in existing_routes:
                    existing_routes[bus_num] = []
                
                if camper.get('location'):
                    existing_routes[bus_num].append({
                        'lat': camper['location']['latitude'],
                        'lng': camper['location']['longitude']
                    })
        
        # Find optimal bus
        camper_address = {
            'lat': new_camper['location']['latitude'],
            'lng': new_camper['location']['longitude']
        }
        
        optimal_bus = route_optimizer.find_optimal_bus(camper_address, existing_routes)
        bus_number_str = f"Bus #{optimal_bus:02d}"
        
        # Update in database
        await db.campers.update_one(
            {"_id": camper_id},
            {"$set": {
                "bus_number": bus_number_str,
                "bus_color": get_bus_color(bus_number_str)
            }}
        )
        
        # Update in CampMinder
        success = await campminder_api.update_camper_bus_assignment(camper_id, optimal_bus)
        
        return {
            "status": "success",
            "camper_id": camper_id,
            "assigned_bus": bus_number_str,
            "synced_to_campminder": success
        }
    except Exception as e:
        logging.error(f"Error auto-assigning camper: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/sync-assignments-to-campminder")
async def sync_assignments_to_campminder():
    """Sync all bus assignments back to CampMinder"""
    try:
        # Get all campers with bus assignments
        campers = await db.campers.find({"bus_number": {"$exists": True}}).to_list(None)
        
        assignments = []
        for camper in campers:
            bus_num_str = camper.get('bus_number', '')
            if bus_num_str:
                bus_num = int(''.join(filter(str.isdigit, bus_num_str)))
                assignments.append({
                    "camper_id": camper['_id'],
                    "bus_number": bus_num
                })
        
        # Bulk update in CampMinder
        results = await campminder_api.bulk_update_bus_assignments(assignments)
        
        successful = sum(1 for v in results.values() if v)
        failed = sum(1 for v in results.values() if not v)
        
        return {
            "status": "success",
            "total": len(assignments),
            "successful": successful,
            "failed": failed
        }
    except Exception as e:
        logging.error(f"Error syncing to CampMinder: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/auto-sync-status")
async def get_auto_sync_status():
    """Get current auto-sync status"""
    sync_status = await db.sync_status.find_one({"_id": "auto_sync"})
    
    return {
        "enabled": AUTO_SYNC_ENABLED,
        "interval_minutes": SYNC_INTERVAL_MINUTES,
        "last_sync": last_sync_time.isoformat() if last_sync_time else None,
        "sync_info": sync_status if sync_status else {}
    }

@api_router.post("/trigger-sync")
async def trigger_manual_sync():
    """Manually trigger a sync with CampMinder (from Google Sheet)"""
    try:
        await auto_sync_campminder()
        return {"status": "success", "message": "Sync from Google Sheet completed"}
    except Exception as e:
        logging.error(f"Error in manual sync: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/test-campminder-api")
async def test_campminder_api():
    """
    Test CampMinder API connectivity and endpoints
    
    Returns comprehensive diagnostic information about API access
    """
    try:
        # Use the built-in test method
        result = await campminder_api.test_api_connectivity()
        return result
    except Exception as e:
        logger.error(f"Error testing CampMinder API: {str(e)}")
        return {"status": "error", "message": str(e)}


@api_router.post("/sync-from-campminder-api")
async def sync_from_campminder_api():
    """
    Sync camper bus data directly from CampMinder API
    
    Uses:
    - Custom Field API for AM/PM bus assignments
    - Day Travel API for transportation assignments
    - Camper Data API for camper information
    """
    try:
        logger.info("Starting sync from CampMinder API")
        
        # First test if API is accessible
        token = await campminder_api.get_jwt_token()
        if not token:
            return {
                "status": "error",
                "message": "CampMinder API authentication failed. Please verify API credentials in .env file.",
                "campers_processed": 0,
                "recommendation": "Use 'Refresh from CSV Now' button to sync from Google Sheet instead."
            }
        
        # Get campers with bus data from CampMinder API
        campers_data = await campminder_api.get_all_campers_with_bus_data(season_id="2026")
        
        if not campers_data:
            return {
                "status": "warning",
                "message": "CampMinder API returned no camper data. The API endpoints may not be available for your subscription level.",
                "campers_processed": 0,
                "recommendation": "Contact CampMinder support to verify API access or use Google Sheet sync as fallback."
            }
        
        new_count = 0
        updated_count = 0
        skipped_count = 0
        
        for camper in campers_data:
            # Skip if no address
            if not camper.get('address'):
                skipped_count += 1
                continue
            
            # Skip if no valid bus assignment
            am_bus = camper.get('am_bus_number', '')
            pm_bus = camper.get('pm_bus_number', '')
            
            if not am_bus and not pm_bus:
                skipped_count += 1
                continue
            
            # Geocode address (using cached version)
            location = await geocode_address_cached(
                camper['address'],
                camper.get('town', ''),
                camper.get('zip_code', '')
            )
            
            if not location or location.latitude == 0:
                skipped_count += 1
                continue
            
            # Generate camper ID
            camper_id = f"{camper['last_name']}_{camper['first_name']}_{camper.get('zip_code', 'NOZIP')}".replace(' ', '_')
            
            # Determine bus color - use AM bus if available, else PM bus
            primary_bus = am_bus if am_bus and am_bus.startswith('Bus') else pm_bus
            bus_color = get_bus_color(primary_bus) if primary_bus else "#808080"
            
            # Determine pickup type based on session
            if camper.get('has_am_session') and camper.get('has_pm_session'):
                pickup_type = "AM & PM"
            elif camper.get('has_am_session'):
                pickup_type = "AM Only"
            elif camper.get('has_pm_session'):
                pickup_type = "PM Only"
            else:
                pickup_type = "Unknown"
            
            camper_doc = {
                "_id": camper_id,
                "first_name": camper['first_name'],
                "last_name": camper['last_name'],
                "session": camper.get('session_type', ''),
                "location": {
                    "latitude": location.latitude,
                    "longitude": location.longitude,
                    "address": location.address
                },
                "town": camper.get('town', ''),
                "zip_code": camper.get('zip_code', ''),
                "pickup_type": pickup_type,
                "am_bus_number": am_bus,  # Empty string if no AM bus/session
                "pm_bus_number": pm_bus,  # Empty string if no PM bus/session
                "bus_color": bus_color,
                "synced_from": "campminder_api",
                "created_at": datetime.now(timezone.utc)
            }
            
            result = await db.campers.replace_one({"_id": camper_id}, camper_doc, upsert=True)
            
            if result.upserted_id:
                new_count += 1
            elif result.modified_count > 0:
                updated_count += 1
        
        # Apply sibling offset
        await apply_sibling_offset(db)
        
        logger.info(f"CampMinder API sync complete: {new_count} new, {updated_count} updated, {skipped_count} skipped")
        
        return {
            "status": "success",
            "message": "Sync from CampMinder API completed",
            "new_campers": new_count,
            "updated_campers": updated_count,
            "skipped_campers": skipped_count,
            "total_processed": len(campers_data)
        }
        
    except Exception as e:
        logger.error(f"Error syncing from CampMinder API: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/sheets/seat-availability")
async def get_seat_availability_for_sheets():
    """Get formatted seat availability data for Google Sheets - COVER SHEET FORMAT"""
    try:
        # Include ALL campers with bus assignments (even without valid locations)
        campers = await db.campers.find({
            "am_bus_number": {"$exists": True, "$nin": ["NONE", ""]}
        }).to_list(None)
        
        # Get staff configurations from database
        staff_configs = await db.bus_staff.find({}).to_list(None)
        staff_dict = {c['bus_number']: c for c in staff_configs}
        
        # Use compact Cover Sheet format with staff info
        sheet_data = cover_sheet_generator.generate_cover_sheet(campers, staff_dict)
        
        return {
            "status": "success",
            "data": sheet_data,
            "last_updated": datetime.now(timezone.utc).isoformat()
        }
    except Exception as e:
        logging.error(f"Error generating sheets data: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/sheets/compact-availability")
async def get_compact_availability():
    """Get compact seat availability summary for Google Sheets"""
    try:
        campers = await db.campers.find({
            "bus_number": {"$exists": True, "$nin": ["NONE", ""]}
        }).to_list(None)
        compact_data = sheets_generator.generate_compact_availability(campers)
        
        return {
            "status": "success",
            "data": compact_data,
            "last_updated": datetime.now(timezone.utc).isoformat()
        }
    except Exception as e:
        logging.error(f"Error generating compact data: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/seat-availability-json")
async def get_seat_availability_json():
    """Get seat availability data as JSON for frontend display"""
    try:
        # Get ALL campers with bus assignments (including those without addresses)
        campers = await db.campers.find({
            "$or": [
                {"am_bus_number": {"$exists": True, "$nin": ["NONE", ""]}},
                {"pm_bus_number": {"$exists": True, "$nin": ["NONE", ""]}}
            ]
        }).to_list(None)
        
        # Get all shadows (they take bus seats too)
        shadows = await db.shadows.find({}).to_list(None)
        
        # Get staff configurations from database
        staff_configs = await db.bus_staff.find({}).to_list(None)
        staff_dict = {c['bus_number']: c for c in staff_configs}
        
        # Group and count by bus
        from collections import defaultdict
        bus_data = defaultdict(lambda: {
            'h1_am': 0, 'h1_pm': 0, 'h2_am': 0, 'h2_pm': 0,
            'shadows': 0,
            'capacity': 30, 'location': '', 'driver': 'TBD', 'counselor': 'TBD'
        })
        
        def parse_session(session):
            """Parse session to determine which halves the camper attends"""
            session_lower = (session or '').lower()
            is_full = 'full season' in session_lower or 'full' in session_lower
            is_half1 = 'half season 1' in session_lower or 'half 1' in session_lower or 'first half' in session_lower
            is_half2 = 'half season 2' in session_lower or 'half 2' in session_lower or 'second half' in session_lower
            is_flex = '6 week' in session_lower or 'flex' in session_lower
            
            # Default to full if no session specified
            if not is_full and not is_half1 and not is_half2 and not is_flex:
                is_full = True
            
            return {
                'h1': is_full or is_half1 or is_flex,
                'h2': is_full or is_half2 or is_flex
            }
        
        # Process each camper
        for camper in campers:
            am_bus = camper.get('am_bus_number', '')
            pm_bus = camper.get('pm_bus_number', '')
            session = camper.get('session', '')
            halves = parse_session(session)
            
            # Count for AM bus
            if am_bus and am_bus != 'NONE' and am_bus.startswith('Bus'):
                if halves['h1']:
                    bus_data[am_bus]['h1_am'] += 1
                if halves['h2']:
                    bus_data[am_bus]['h2_am'] += 1
            
            # Count for PM bus
            if pm_bus and pm_bus != 'NONE' and pm_bus.startswith('Bus'):
                if halves['h1']:
                    bus_data[pm_bus]['h1_pm'] += 1
                if halves['h2']:
                    bus_data[pm_bus]['h2_pm'] += 1
        
        # Process shadows - they inherit the session from their linked camper
        for shadow in shadows:
            bus_number = shadow.get('bus_number', '')
            session = shadow.get('session', '')
            halves = parse_session(session)
            
            if bus_number and bus_number.startswith('Bus'):
                bus_data[bus_number]['shadows'] += 1
                # Shadows take both AM and PM seats (same as their linked camper)
                if halves['h1']:
                    bus_data[bus_number]['h1_am'] += 1
                    bus_data[bus_number]['h1_pm'] += 1
                if halves['h2']:
                    bus_data[bus_number]['h2_am'] += 1
                    bus_data[bus_number]['h2_pm'] += 1
        
        # Add capacity and staff info
        result = {}
        for bus_number in bus_data:
            data = bus_data[bus_number]
            
            # Get staff info
            if bus_number in staff_dict:
                staff = staff_dict[bus_number]
                data['capacity'] = staff.get('capacity', get_bus_capacity(bus_number))
                data['location'] = staff.get('location_name', get_bus_location(bus_number))
                data['driver'] = staff.get('driver_name', get_bus_driver(bus_number))
                data['counselor'] = staff.get('counselor_name', get_bus_counselor(bus_number))
            else:
                data['capacity'] = get_bus_capacity(bus_number)
                data['location'] = get_bus_location(bus_number)
                data['driver'] = get_bus_driver(bus_number)
                data['counselor'] = get_bus_counselor(bus_number)
            
            # Calculate available seats
            cap = data['capacity']
            data['h1_am_available'] = cap - data['h1_am']
            data['h1_pm_available'] = cap - data['h1_pm']
            data['h2_am_available'] = cap - data['h2_am']
            data['h2_pm_available'] = cap - data['h2_pm']
            
            result[bus_number] = data
        
        return {
            "status": "success",
            "buses": result
        }
    except Exception as e:
        logging.error(f"Error getting seat availability JSON: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/download/seat-availability")
async def download_seat_availability():
    """Download seat availability as formatted Excel file matching the Google Sheet"""
    from fastapi.responses import Response
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    
    try:
        # Get all campers with bus assignments
        campers = await db.campers.find({
            "$or": [
                {"am_bus_number": {"$exists": True, "$nin": ["NONE", ""]}},
                {"pm_bus_number": {"$exists": True, "$nin": ["NONE", ""]}}
            ]
        }).to_list(None)
        
        # Get staff configurations from database
        staff_configs = await db.bus_staff.find({}).to_list(None)
        staff_dict = {c['bus_number']: c for c in staff_configs}
        
        # Generate cover sheet data with staff info
        sheet_data = cover_sheet_generator.generate_cover_sheet(campers, staff_dict)
        
        # Create Excel workbook with formatting
        wb = Workbook()
        ws = wb.active
        ws.title = "Seat Availability"
        
        # Define styles
        title_font = Font(name='Arial', size=16, bold=True, color='1F4E79')
        subtitle_font = Font(name='Arial', size=12, bold=True, color='1F4E79')
        header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        data_font = Font(name='Arial', size=10)
        totals_font = Font(name='Arial', size=11, bold=True)
        totals_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        
        # Border styles
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Alternating row colors
        light_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        
        # Available column colors based on seat count
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        orange_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        green_font = Font(name='Arial', size=10, bold=True, color='006100')
        orange_font = Font(name='Arial', size=10, bold=True, color='9C5700')
        red_font = Font(name='Arial', size=10, bold=True, color='9C0006')
        
        # Available column indices (1-indexed): 7 (H1 AM Avail), 9 (H1 PM Avail), 11 (H2 AM Avail), 13 (H2 PM Avail)
        avail_cols = [7, 9, 11, 13]  # 1-indexed, removed column 14 (Available)
        
        # Write data with formatting
        for row_idx, row_data in enumerate(sheet_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Title row (row 1)
                if row_idx == 1:
                    cell.font = title_font
                    cell.alignment = Alignment(horizontal='left')
                # Subtitle row (row 2)
                elif row_idx == 2:
                    cell.font = subtitle_font
                    cell.alignment = Alignment(horizontal='left')
                # Header row (row 4)
                elif row_idx == 4:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                # TOTALS row
                elif row_data and row_data[0] == 'TOTALS':
                    cell.font = totals_font
                    cell.fill = totals_fill
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                # AVAILABLE SEATS header
                elif row_data and 'AVAILABLE SEATS' in str(row_data[0]):
                    cell.font = subtitle_font
                # Data rows (bus data)
                elif row_idx > 4 and row_data and str(row_data[0]).startswith('Bus'):
                    cell.font = data_font
                    cell.border = thin_border
                    # Alternating row colors
                    if (row_idx - 5) % 2 == 0:
                        cell.fill = light_fill
                    else:
                        cell.fill = white_fill
                    
                    # Center align numeric columns
                    if col_idx >= 5:
                        cell.alignment = Alignment(horizontal='center')
                    
                    # Color the Available columns based on value
                    if col_idx in avail_cols:
                        try:
                            avail = int(value) if value is not None else 0
                            if avail < 5:
                                cell.fill = red_fill
                                cell.font = red_font
                            elif avail <= 10:
                                cell.fill = orange_fill
                                cell.font = orange_font
                            else:
                                cell.fill = green_fill
                                cell.font = green_font
                        except:
                            pass
                # Summary rows at bottom
                elif row_data and ('Half' in str(row_data[0]) or 'Available' in str(row_data[0])):
                    cell.font = data_font
        
        # Set column widths for 13 columns
        column_widths = [10, 18, 16, 16, 7, 10, 10, 10, 10, 10, 10, 10, 10]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Freeze the header row
        ws.freeze_panes = 'A5'
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"seat_availability_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
                "Cache-Control": "no-cache, no-store, must-revalidate",
                "Pragma": "no-cache",
                "Expires": "0",
                "Access-Control-Expose-Headers": "Content-Disposition"
            }
        )
    except Exception as e:
        logging.error(f"Error generating seat availability: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/update-seat-availability-sheet")
async def update_seat_availability_sheet():
    """
    Update the seat availability Google Sheet (1ZK58gjF4BO0HF_2y6oovrjzRH3qV5zAs8H-7CeKOSGE)
    with current bus assignments.
    Uses 14-column format with availability columns.
    """
    try:
        # Get all campers with bus assignments
        campers = await db.campers.find({
            "am_bus_number": {"$exists": True, "$nin": ["NONE", ""]}
        }).to_list(None)
        
        # Get staff configurations from database
        staff_configs = await db.bus_staff.find({}).to_list(None)
        staff_dict = {c['bus_number']: c for c in staff_configs}
        
        # Generate cover sheet data in 14-column format with availability columns
        sheet_data = cover_sheet_generator.generate_cover_sheet(campers, staff_dict)
        
        # Use dedicated seat availability webhook
        webhook_url = os.environ.get('SEAT_AVAILABILITY_WEBHOOK_URL', '')
        if not webhook_url:
            return {
                "status": "error",
                "message": "SEAT_AVAILABILITY_WEBHOOK_URL not configured"
            }
        
        payload = {
            "action": "update_seat_availability",
            "sheet_id": OUTPUT_SHEET_ID,
            "data": sheet_data
        }
        
        async with httpx.AsyncClient(timeout=60.0, follow_redirects=True) as client:
            response = await client.post(webhook_url, json=payload)
            
            if response.status_code == 200:
                return {
                    "status": "success",
                    "message": f"Updated seat availability sheet with {len(sheet_data)} rows",
                    "sheet_url": f"https://docs.google.com/spreadsheets/d/{OUTPUT_SHEET_ID}/edit"
                }
            else:
                return {
                    "status": "error",
                    "message": f"Webhook returned status {response.status_code}",
                    "response": response.text
                }
    except Exception as e:
        logging.error(f"Error updating seat availability sheet: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/push-seat-availability-to-sheet")
async def push_seat_availability_to_sheet():
    """
    Push current seat availability data to Google Sheet via webhook.
    This is the button-triggered version that shows detailed status.
    Uses 14-column format with availability columns.
    """
    try:
        # Get all campers with bus assignments
        all_campers = await db.campers.find({}).to_list(None)
        
        # Filter to campers with valid bus assignments
        campers_with_buses = [c for c in all_campers if 
            (c.get('am_bus_number', '') and c.get('am_bus_number', '') != 'NONE' and c.get('am_bus_number', '').startswith('Bus')) or
            (c.get('pm_bus_number', '') and c.get('pm_bus_number', '') != 'NONE' and c.get('pm_bus_number', '').startswith('Bus'))
        ]
        
        # Get staff configurations from database
        staff_configs = await db.bus_staff.find({}).to_list(None)
        staff_dict = {c['bus_number']: c for c in staff_configs}
        
        # Generate cover sheet data in 14-column format with availability columns
        sheet_data = cover_sheet_generator.generate_cover_sheet(campers_with_buses, staff_dict)
        
        # Use dedicated seat availability webhook
        webhook_url = os.environ.get('SEAT_AVAILABILITY_WEBHOOK_URL', '')
        if not webhook_url:
            return {
                "status": "error",
                "message": "SEAT_AVAILABILITY_WEBHOOK_URL not configured. Please set up the webhook."
            }
        
        payload = {
            "action": "update_seat_availability",
            "sheet_id": OUTPUT_SHEET_ID,
            "data": sheet_data
        }
        
        logging.info(f"Pushing {len(sheet_data)} rows to seat availability sheet (14 columns)")
        
        async with httpx.AsyncClient(timeout=60.0, follow_redirects=True) as client:
            response = await client.post(webhook_url, json=payload)
            response_text = response.text
            
            logging.info(f"Webhook response: {response.status_code} - {response_text}")
            
            if response.status_code == 200:
                return {
                    "status": "success",
                    "message": f"✓ Updated seat availability sheet with {len(sheet_data)} rows ({len(campers_with_buses)} campers)",
                    "sheet_url": f"https://docs.google.com/spreadsheets/d/{OUTPUT_SHEET_ID}/edit",
                    "rows_updated": len(sheet_data)
                }
            else:
                return {
                    "status": "error",
                    "message": f"Webhook returned status {response.status_code}: {response_text}",
                    "response": response_text
                }
    except Exception as e:
        logging.error(f"Error pushing seat availability: {str(e)}")
        return {
            "status": "error", 
            "message": f"Error: {str(e)}"
        }


# ============================================
# BUS STAFF CONFIGURATION ENDPOINTS
# ============================================

class BusStaffConfig(BaseModel):
    """Model for bus staff configuration"""
    bus_number: str
    driver_name: str
    counselor_name: str
    home_address: str
    capacity: Optional[int] = None
    location_name: Optional[str] = None


@api_router.get("/bus-staff")
async def get_all_bus_staff():
    """Get all bus staff configurations from database"""
    try:
        staff_configs = await db.bus_staff.find({}).to_list(None)
        
        # Convert to dict format
        result = {}
        for config in staff_configs:
            bus_num = config.get('bus_number', '')
            result[bus_num] = {
                'bus_number': bus_num,
                'driver_name': config.get('driver_name', 'TBD'),
                'counselor_name': config.get('counselor_name', 'TBD'),
                'home_address': config.get('home_address', ''),
                'capacity': config.get('capacity', get_bus_capacity(bus_num)),
                'location_name': config.get('location_name', get_bus_location(bus_num)),
                'lat': config.get('lat'),
                'lng': config.get('lng'),
                'last_updated': config.get('last_updated')
            }
        
        return {
            "status": "success",
            "staff": result,
            "count": len(result)
        }
    except Exception as e:
        logging.error(f"Error getting bus staff: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/bus-staff/{bus_number}")
async def get_bus_staff(bus_number: str):
    """Get staff configuration for a specific bus"""
    import urllib.parse
    try:
        decoded_bus = urllib.parse.unquote(bus_number)
        config = await db.bus_staff.find_one({"bus_number": decoded_bus})
        
        if config:
            return {
                "status": "success",
                "bus_number": decoded_bus,
                "driver_name": config.get('driver_name', 'TBD'),
                "counselor_name": config.get('counselor_name', 'TBD'),
                "home_address": config.get('home_address', ''),
                "capacity": config.get('capacity', get_bus_capacity(decoded_bus)),
                "location_name": config.get('location_name', get_bus_location(decoded_bus)),
                "lat": config.get('lat'),
                "lng": config.get('lng')
            }
        else:
            # Return defaults from bus_config
            return {
                "status": "success",
                "bus_number": decoded_bus,
                "driver_name": get_bus_driver(decoded_bus),
                "counselor_name": get_bus_counselor(decoded_bus),
                "home_address": get_bus_home_location(decoded_bus),
                "capacity": get_bus_capacity(decoded_bus),
                "location_name": get_bus_location(decoded_bus),
                "lat": None,
                "lng": None
            }
    except Exception as e:
        logging.error(f"Error getting bus staff: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/bus-staff")
async def save_bus_staff(config: BusStaffConfig):
    """Save or update bus staff configuration"""
    try:
        # Geocode the address if provided (using cached version)
        lat = None
        lng = None
        if config.home_address:
            location = await geocode_address_cached(config.home_address, "", "")
            if location:
                lat = location.latitude
                lng = location.longitude
                logging.info(f"Geocoded {config.home_address} to {lat}, {lng}")
        
        # Prepare document
        staff_doc = {
            "bus_number": config.bus_number,
            "driver_name": config.driver_name,
            "counselor_name": config.counselor_name,
            "home_address": config.home_address,
            "capacity": config.capacity or get_bus_capacity(config.bus_number),
            "location_name": config.location_name or get_bus_location(config.bus_number),
            "lat": lat,
            "lng": lng,
            "last_updated": datetime.now(timezone.utc).isoformat()
        }
        
        # Upsert to database
        result = await db.bus_staff.replace_one(
            {"bus_number": config.bus_number},
            staff_doc,
            upsert=True
        )
        
        logging.info(f"Saved staff config for {config.bus_number}: Driver={config.driver_name}, Counselor={config.counselor_name}")
        
        return {
            "status": "success",
            "message": f"Saved configuration for {config.bus_number}",
            "bus_number": config.bus_number,
            "driver_name": config.driver_name,
            "counselor_name": config.counselor_name,
            "was_update": result.modified_count > 0
        }
    except Exception as e:
        logging.error(f"Error saving bus staff: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.delete("/bus-staff/{bus_number}")
async def delete_bus_staff(bus_number: str):
    """Delete bus staff configuration"""
    import urllib.parse
    try:
        decoded_bus = urllib.parse.unquote(bus_number)
        result = await db.bus_staff.delete_one({"bus_number": decoded_bus})
        
        if result.deleted_count > 0:
            return {
                "status": "success",
                "message": f"Deleted configuration for {decoded_bus}"
            }
        else:
            raise HTTPException(status_code=404, detail="Bus staff configuration not found")
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error deleting bus staff: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ============================================
# SHADOW STAFF ENDPOINTS (1:1 staff for campers)
# ============================================

class ShadowCreate(BaseModel):
    """Model for creating a shadow staff member"""
    shadow_name: str
    camper_id: str  # The camper this shadow is linked to
    bus_number: Optional[str] = None  # Optionally specify the bus (for AM/PM specific stops)

class ShadowUpdate(BaseModel):
    """Model for updating a shadow"""
    shadow_name: Optional[str] = None
    camper_id: Optional[str] = None

@api_router.get("/shadows")
async def get_all_shadows():
    """Get all shadow staff members"""
    try:
        shadows = await db.shadows.find({}).to_list(None)
        result = []
        for shadow in shadows:
            result.append({
                "id": str(shadow.get("_id", "")),
                "shadow_name": shadow.get("shadow_name"),
                "camper_id": shadow.get("camper_id"),
                "camper_name": shadow.get("camper_name"),
                "bus_number": shadow.get("bus_number"),
                "session": shadow.get("session"),
                "town": shadow.get("town", ""),
                "created_at": shadow.get("created_at"),
                "updated_at": shadow.get("updated_at")
            })
        return {"status": "success", "shadows": result, "count": len(result)}
    except Exception as e:
        logging.error(f"Error getting shadows: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/shadows/by-bus/{bus_number}")
async def get_shadows_by_bus(bus_number: str):
    """Get all shadows assigned to campers on a specific bus"""
    import urllib.parse
    try:
        decoded_bus = urllib.parse.unquote(bus_number)
        shadows = await db.shadows.find({"bus_number": decoded_bus}).to_list(None)
        result = []
        for shadow in shadows:
            result.append({
                "id": str(shadow.get("_id", "")),
                "shadow_name": shadow.get("shadow_name"),
                "camper_id": shadow.get("camper_id"),
                "camper_name": shadow.get("camper_name"),
                "bus_number": shadow.get("bus_number"),
                "session": shadow.get("session"),
                "created_at": shadow.get("created_at")
            })
        return {"status": "success", "shadows": result, "count": len(result)}
    except Exception as e:
        logging.error(f"Error getting shadows by bus: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/shadows/by-camper/{camper_id}")
async def get_shadow_by_camper(camper_id: str):
    """Get shadow for a specific camper"""
    import urllib.parse
    try:
        decoded_id = urllib.parse.unquote(camper_id)
        shadow = await db.shadows.find_one({"camper_id": decoded_id})
        if shadow:
            return {
                "status": "success",
                "shadow": {
                    "id": str(shadow.get("_id", "")),
                    "shadow_name": shadow.get("shadow_name"),
                    "camper_id": shadow.get("camper_id"),
                    "camper_name": shadow.get("camper_name"),
                    "bus_number": shadow.get("bus_number"),
                    "session": shadow.get("session")
                }
            }
        return {"status": "success", "shadow": None}
    except Exception as e:
        logging.error(f"Error getting shadow by camper: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/shadows")
async def create_shadow(shadow_data: ShadowCreate):
    """Create a new shadow staff member linked to a camper"""
    try:
        # Get the camper to inherit their session
        camper = await db.campers.find_one({"_id": shadow_data.camper_id})
        if not camper:
            raise HTTPException(status_code=404, detail=f"Camper not found: {shadow_data.camper_id}")
        
        # Check if shadow already exists for this camper
        existing = await db.shadows.find_one({"camper_id": shadow_data.camper_id})
        if existing:
            raise HTTPException(
                status_code=400,
                detail=f"Shadow already exists for this camper. Use PUT to update."
            )
        
        # Use provided bus_number if specified, otherwise infer from camper
        if shadow_data.bus_number:
            bus_number = shadow_data.bus_number
        else:
            # Fallback: Determine which bus to use (prefer AM bus)
            bus_number = camper.get('am_bus_number', '')
            if not bus_number or bus_number == 'NONE':
                bus_number = camper.get('pm_bus_number', '')
        
        # Create shadow document
        shadow_doc = {
            "shadow_name": shadow_data.shadow_name.strip(),
            "camper_id": shadow_data.camper_id,
            "camper_name": f"{camper.get('first_name', '')} {camper.get('last_name', '')}".strip(),
            "bus_number": bus_number,
            "session": camper.get('session', 'Full Season- 5 Days'),
            "town": camper.get('town', ''),  # Store town for reference
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        
        result = await db.shadows.insert_one(shadow_doc)
        shadow_doc["id"] = str(result.inserted_id)
        if "_id" in shadow_doc:
            del shadow_doc["_id"]
        
        logging.info(f"Created shadow '{shadow_data.shadow_name}' for camper {shadow_doc['camper_name']} on {bus_number}")
        return {"status": "success", "shadow": shadow_doc}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error creating shadow: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.put("/shadows/{shadow_id}")
async def update_shadow(shadow_id: str, shadow_data: ShadowUpdate):
    """Update an existing shadow"""
    from bson import ObjectId
    try:
        update_fields = {"updated_at": datetime.now(timezone.utc).isoformat()}
        
        if shadow_data.shadow_name is not None:
            update_fields["shadow_name"] = shadow_data.shadow_name.strip()
        
        if shadow_data.camper_id is not None:
            # Get new camper info
            camper = await db.campers.find_one({"_id": shadow_data.camper_id})
            if not camper:
                raise HTTPException(status_code=404, detail=f"Camper not found: {shadow_data.camper_id}")
            
            update_fields["camper_id"] = shadow_data.camper_id
            update_fields["camper_name"] = f"{camper.get('first_name', '')} {camper.get('last_name', '')}".strip()
            update_fields["session"] = camper.get('session', 'Full Season- 5 Days')
            
            # Update bus number
            bus_number = camper.get('am_bus_number', '')
            if not bus_number or bus_number == 'NONE':
                bus_number = camper.get('pm_bus_number', '')
            update_fields["bus_number"] = bus_number
        
        result = await db.shadows.update_one(
            {"_id": ObjectId(shadow_id)},
            {"$set": update_fields}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Shadow not found")
        
        return {"status": "success", "message": "Shadow updated"}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error updating shadow: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.delete("/shadows/{shadow_id}")
async def delete_shadow(shadow_id: str):
    """Delete a shadow staff member"""
    from bson import ObjectId
    try:
        result = await db.shadows.delete_one({"_id": ObjectId(shadow_id)})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Shadow not found")
        
        logging.info(f"Deleted shadow: {shadow_id}")
        return {"status": "success", "message": "Shadow deleted"}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error deleting shadow: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.delete("/shadows/by-camper/{camper_id}")
async def delete_shadow_by_camper(camper_id: str):
    """Delete shadow by camper ID"""
    import urllib.parse
    try:
        decoded_id = urllib.parse.unquote(camper_id)
        result = await db.shadows.delete_one({"camper_id": decoded_id})
        if result.deleted_count == 0:
            return {"status": "success", "message": "No shadow found for this camper"}
        
        logging.info(f"Deleted shadow for camper: {decoded_id}")
        return {"status": "success", "message": "Shadow deleted"}
    except Exception as e:
        logging.error(f"Error deleting shadow by camper: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ============================================
# BUS ZONES ENDPOINTS (User-defined zones)
# ============================================

class ZonePoint(BaseModel):
    lat: float
    lng: float

class BusZoneCreate(BaseModel):
    bus_number: str
    points: List[ZonePoint]
    name: Optional[str] = None
    color: Optional[str] = None

class BusZoneUpdate(BaseModel):
    points: Optional[List[ZonePoint]] = None
    name: Optional[str] = None
    color: Optional[str] = None

@api_router.get("/bus-zones")
async def get_bus_zones():
    """Get all user-defined bus zones"""
    try:
        zones = await db.bus_zones.find({}).to_list(None)
        # Convert ObjectId to string and return
        result = []
        for zone in zones:
            result.append({
                "id": str(zone.get("_id", "")),
                "bus_number": zone.get("bus_number"),
                "points": zone.get("points", []),
                "name": zone.get("name", ""),
                "color": zone.get("color", ""),
                "created_at": zone.get("created_at"),
                "updated_at": zone.get("updated_at")
            })
        return {"zones": result}
    except Exception as e:
        logging.error(f"Error getting bus zones: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/bus-zones/{bus_number}")
async def get_bus_zone(bus_number: str):
    """Get zone for a specific bus"""
    try:
        zone = await db.bus_zones.find_one({"bus_number": bus_number})
        if not zone:
            return {"zone": None}
        return {
            "zone": {
                "id": str(zone.get("_id", "")),
                "bus_number": zone.get("bus_number"),
                "points": zone.get("points", []),
                "name": zone.get("name", ""),
                "color": zone.get("color", ""),
                "created_at": zone.get("created_at"),
                "updated_at": zone.get("updated_at")
            }
        }
    except Exception as e:
        logging.error(f"Error getting bus zone: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/bus-zones")
async def create_bus_zone(zone_data: BusZoneCreate):
    """Create a new bus zone (one zone per bus)"""
    try:
        # Check if zone already exists for this bus
        existing = await db.bus_zones.find_one({"bus_number": zone_data.bus_number})
        if existing:
            raise HTTPException(
                status_code=400, 
                detail=f"Zone already exists for {zone_data.bus_number}. Use PUT to update."
            )
        
        # Create new zone
        zone_doc = {
            "bus_number": zone_data.bus_number,
            "points": [{"lat": p.lat, "lng": p.lng} for p in zone_data.points],
            "name": zone_data.name or f"{zone_data.bus_number} Zone",
            "color": zone_data.color or "",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        
        result = await db.bus_zones.insert_one(zone_doc)
        zone_doc["id"] = str(result.inserted_id)
        if "_id" in zone_doc:
            del zone_doc["_id"]
        
        return {"status": "success", "zone": zone_doc}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error creating bus zone: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.put("/bus-zones/{bus_number}")
async def update_bus_zone(bus_number: str, zone_data: BusZoneUpdate):
    """Update an existing bus zone"""
    try:
        # Find existing zone
        existing = await db.bus_zones.find_one({"bus_number": bus_number})
        if not existing:
            raise HTTPException(status_code=404, detail=f"No zone found for {bus_number}")
        
        # Build update document
        update_doc = {"updated_at": datetime.now(timezone.utc).isoformat()}
        
        if zone_data.points is not None:
            update_doc["points"] = [{"lat": p.lat, "lng": p.lng} for p in zone_data.points]
        if zone_data.name is not None:
            update_doc["name"] = zone_data.name
        if zone_data.color is not None:
            update_doc["color"] = zone_data.color
        
        await db.bus_zones.update_one(
            {"bus_number": bus_number},
            {"$set": update_doc}
        )
        
        # Fetch updated zone
        updated = await db.bus_zones.find_one({"bus_number": bus_number})
        return {
            "status": "success",
            "zone": {
                "id": str(updated.get("_id", "")),
                "bus_number": updated.get("bus_number"),
                "points": updated.get("points", []),
                "name": updated.get("name", ""),
                "color": updated.get("color", ""),
                "created_at": updated.get("created_at"),
                "updated_at": updated.get("updated_at")
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error updating bus zone: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.delete("/bus-zones/{bus_number}")
async def delete_bus_zone(bus_number: str):
    """Delete a bus zone"""
    try:
        result = await db.bus_zones.delete_one({"bus_number": bus_number})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail=f"No zone found for {bus_number}")
        return {"status": "success", "message": f"Zone for {bus_number} deleted"}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error deleting bus zone: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/buses")
async def get_buses():
    """Get all buses with their info including home locations and staff"""
    try:
        # Get staff configs from database
        staff_configs = await db.bus_staff.find({}).to_list(None)
        staff_dict = {c['bus_number']: c for c in staff_configs}
        
        buses = []
        for bus_number in get_all_buses():
            bus_info = get_bus_info(bus_number)
            
            # Override with database values if available
            if bus_number in staff_dict:
                db_config = staff_dict[bus_number]
                bus_info['driver'] = db_config.get('driver_name', bus_info.get('driver', 'TBD'))
                bus_info['counselor'] = db_config.get('counselor_name', bus_info.get('counselor', 'TBD'))
                bus_info['home_location'] = db_config.get('home_address', bus_info.get('home_location', ''))
                if db_config.get('capacity'):
                    bus_info['capacity'] = db_config['capacity']
            
            # Get camper count for this bus
            am_count = await db.campers.count_documents({"am_bus_number": bus_number})
            pm_count = await db.campers.count_documents({"pm_bus_number": bus_number})
            bus_info['am_camper_count'] = am_count
            bus_info['pm_camper_count'] = pm_count
            buses.append(bus_info)
        
        return {
            "status": "success",
            "buses": buses,
            "camp_address": get_camp_address()
        }
    except Exception as e:
        logging.error(f"Error getting buses: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/buses/{bus_number}")
async def get_bus_details(bus_number: str):
    """Get detailed info for a specific bus"""
    try:
        bus_info = get_bus_info(bus_number)
        
        # Get campers for this bus
        am_campers = await db.campers.find({"am_bus_number": bus_number}).to_list(None)
        pm_campers = await db.campers.find({"pm_bus_number": bus_number}).to_list(None)
        
        return {
            "status": "success",
            "bus": bus_info,
            "camp_address": get_camp_address(),
            "am_campers": [
                {"name": f"{c['first_name']} {c['last_name']}", "address": c.get('location', {}).get('address', '')}
                for c in am_campers
            ],
            "pm_campers": [
                {"name": f"{c['first_name']} {c['last_name']}", "address": c.get('location', {}).get('address', '')}
                for c in pm_campers
            ]
        }
    except Exception as e:
        logging.error(f"Error getting bus details: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/audit/campers")
async def audit_all_campers():
    """
    Comprehensive audit of ALL campers to verify bus assignments are correct.
    Compares database values against Google Sheet source data.
    
    Distinguishes between:
    - TRUE ERRORS: Database has different bus than Sheet (Sheet has valid bus)
    - AUTO-ASSIGNMENTS: Database has bus, Sheet has NONE (system auto-assigned)
    """
    import httpx
    import csv
    from io import StringIO
    
    logger.info("=== STARTING COMPREHENSIVE CAMPER AUDIT ===")
    
    results = {
        "status": "success",
        "total_checked": 0,
        "true_errors": [],         # DB differs from valid sheet bus
        "auto_assignments": [],     # DB has bus, sheet has NONE
        "summary": {}
    }
    
    try:
        # Step 1: Load all campers from database
        db_campers = await db.campers.find({}).to_list(None)
        logger.info(f"Loaded {len(db_campers)} campers from database")
        
        # Step 2: Load Google Sheet data for comparison
        sheet_id = CAMPMINDER_SHEET_ID
        csv_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv"
        
        async with httpx.AsyncClient(timeout=60.0, follow_redirects=True) as client:
            response = await client.get(csv_url)
            csv_content = response.text
        
        # Parse Google Sheet data
        if csv_content.startswith('\ufeff'):
            csv_content = csv_content[1:]
        
        reader = csv.DictReader(StringIO(csv_content))
        sheet_data = {}
        
        for row in reader:
            first_name = row.get('First Name', '').strip()
            last_name = row.get('Last Name', '').strip()
            am_bus = row.get('2026Transportation M AM Bus', '').strip()
            pm_bus = row.get('2026Transportation M PM Bus', '').strip()
            
            if first_name and last_name:
                key = f"{last_name}_{first_name}".lower()
                sheet_data[key] = {
                    'name': f"{first_name} {last_name}",
                    'sheet_am_bus': am_bus,
                    'sheet_pm_bus': pm_bus
                }
        
        logger.info(f"Loaded {len(sheet_data)} campers from Google Sheet")
        
        # Step 3: Audit each camper
        seen_campers = set()
        
        for camper in db_campers:
            first_name = camper.get('first_name', '')
            last_name = camper.get('last_name', '')
            camper_id = camper.get('_id', '')
            db_am_bus = camper.get('am_bus_number', '')
            db_pm_bus = camper.get('pm_bus_number', '')
            
            # Skip PM-specific entries
            if camper_id.endswith('_PM'):
                continue
            
            # Skip if already checked
            full_name = f"{first_name} {last_name}"
            if full_name in seen_campers:
                continue
            seen_campers.add(full_name)
            
            results["total_checked"] += 1
            
            # Find in sheet data
            key = f"{last_name}_{first_name}".lower()
            sheet_camper = sheet_data.get(key)
            
            if not sheet_camper:
                continue
            
            sheet_am = sheet_camper['sheet_am_bus']
            sheet_pm = sheet_camper['sheet_pm_bus']
            
            # Determine if sheet value is valid bus
            def is_valid_sheet_bus(val):
                if not val:
                    return False
                val_upper = val.upper()
                if val_upper == 'NONE' or val_upper == '':
                    return False
                if any(x in val_upper for x in ['MAIN TENT', 'HOCKEY RINK', 'AUDITORIUM']):
                    return False
                return val.startswith('Bus')
            
            # Check AM
            if db_am_bus and db_am_bus != 'NONE' and db_am_bus.startswith('Bus'):
                if is_valid_sheet_bus(sheet_am):
                    # Both have valid buses - check if they match
                    db_norm = db_am_bus.replace(' ', '')
                    sheet_norm = sheet_am.replace(' ', '')
                    if db_norm != sheet_norm:
                        results["true_errors"].append({
                            "camper": full_name,
                            "type": "AM",
                            "database_value": db_am_bus,
                            "sheet_value": sheet_am,
                            "issue": "TRUE ERROR: AM bus mismatch"
                        })
                else:
                    # DB has bus, sheet has NONE - auto-assignment
                    results["auto_assignments"].append({
                        "camper": full_name,
                        "type": "AM",
                        "auto_assigned_bus": db_am_bus,
                        "sheet_value": sheet_am or "NONE"
                    })
            
            # Check PM
            if db_pm_bus and db_pm_bus != 'NONE' and db_pm_bus.startswith('Bus'):
                if is_valid_sheet_bus(sheet_pm):
                    # Both have valid buses - check if they match
                    db_norm = db_pm_bus.replace(' ', '')
                    sheet_norm = sheet_pm.replace(' ', '')
                    if db_norm != sheet_norm:
                        results["true_errors"].append({
                            "camper": full_name,
                            "type": "PM",
                            "database_value": db_pm_bus,
                            "sheet_value": sheet_pm,
                            "issue": "TRUE ERROR: PM bus mismatch"
                        })
                else:
                    # DB has bus, sheet has NONE - auto-assignment
                    results["auto_assignments"].append({
                        "camper": full_name,
                        "type": "PM",
                        "auto_assigned_bus": db_pm_bus,
                        "sheet_value": sheet_pm or "NONE"
                    })
        
        # Step 4: Generate summary
        results["summary"] = {
            "total_campers_checked": results["total_checked"],
            "true_errors_count": len(results["true_errors"]),
            "auto_assignments_count": len(results["auto_assignments"]),
            "validation_passed": len(results["true_errors"]) == 0,
            "message": "✓✓✓ ALL BUS LABELS MATCH GOOGLE SHEET" if len(results["true_errors"]) == 0 else f"❌ Found {len(results['true_errors'])} bus mismatches"
        }
        
        if len(results["true_errors"]) > 0:
            results["status"] = "errors_found"
            logger.error(f"AUDIT FOUND {len(results['true_errors'])} TRUE ERRORS")
        else:
            logger.info(f"AUDIT PASSED - {len(results['auto_assignments'])} auto-assignments detected (expected)")
        
        return results
        
    except Exception as e:
        logger.error(f"Error during audit: {str(e)}")
        results["status"] = "error"
        results["message"] = str(e)
        return results


@api_router.get("/audit/bus/{bus_number}")
async def audit_single_bus(bus_number: str):
    """Audit all campers on a specific bus"""
    
    # Get all campers assigned to this bus
    campers = await db.campers.find({
        "$or": [
            {"am_bus_number": bus_number},
            {"pm_bus_number": bus_number}
        ]
    }).to_list(None)
    
    results = {
        "bus_number": bus_number,
        "am_campers": [],
        "pm_campers": [],
        "am_count": 0,
        "pm_count": 0
    }
    
    seen_am = set()
    seen_pm = set()
    
    for camper in campers:
        name = f"{camper['first_name']} {camper['last_name']}"
        camper_id = camper.get('_id', '')
        
        # Check AM assignment
        if camper.get('am_bus_number') == bus_number and name not in seen_am:
            if not camper_id.endswith('_PM'):
                results["am_campers"].append({
                    "name": name,
                    "address": camper.get('location', {}).get('address', ''),
                    "am_bus": camper.get('am_bus_number', ''),
                    "pm_bus": camper.get('pm_bus_number', '')
                })
                seen_am.add(name)
        
        # Check PM assignment
        if camper.get('pm_bus_number') == bus_number and name not in seen_pm:
            results["pm_campers"].append({
                "name": name,
                "address": camper.get('location', {}).get('address', ''),
                "am_bus": camper.get('am_bus_number', ''),
                "pm_bus": camper.get('pm_bus_number', '')
            })
            seen_pm.add(name)
    
    results["am_count"] = len(results["am_campers"])
    results["pm_count"] = len(results["pm_campers"])
    
    return results


@api_router.post("/update-output-sheet")
async def update_output_google_sheet():
    """
    Update the output Google Sheets document with all camper bus assignments.
    Sheet ID: 1ZK58gjF4BO0HF_2y6oovrjzRH3qV5zAs8H-7CeKOSGE
    """
    import httpx
    
    logger.info("=== UPDATING OUTPUT GOOGLE SHEET ===")
    logger.info(f"Sheet ID: {OUTPUT_SHEET_ID}")
    logger.info(f"Sheet URL: https://docs.google.com/spreadsheets/d/{OUTPUT_SHEET_ID}/edit")
    
    try:
        # Get all campers from database
        all_campers = await db.campers.find({}).to_list(None)
        logger.info(f"Found {len(all_campers)} campers in database")
        
        # Prepare data for sheet
        # Format: Name, Address, Town, Zip, Session Type, AM Bus, PM Bus
        sheet_data = []
        seen_campers = set()
        
        for camper in all_campers:
            first_name = camper.get('first_name', '')
            last_name = camper.get('last_name', '')
            camper_id = camper.get('_id', '')
            
            # Skip PM-specific entries (we'll combine data)
            if camper_id.endswith('_PM'):
                continue
            
            full_name = f"{first_name} {last_name}"
            if full_name in seen_campers:
                continue
            seen_campers.add(full_name)
            
            address = camper.get('location', {}).get('address', '')
            town = camper.get('town', '')
            zip_code = camper.get('zip_code', '')
            session = camper.get('session', camper.get('pickup_type', ''))
            am_bus = camper.get('am_bus_number', '')
            pm_bus = camper.get('pm_bus_number', '')
            
            # Clean up bus values - don't show NONE
            if am_bus == 'NONE':
                am_bus = ''
            if pm_bus == 'NONE':
                pm_bus = ''
            
            sheet_data.append({
                'name': full_name,
                'first_name': first_name,
                'last_name': last_name,
                'address': address,
                'town': town,
                'zip': zip_code,
                'session': session,
                'am_bus': am_bus,
                'pm_bus': pm_bus
            })
        
        logger.info(f"Prepared {len(sheet_data)} unique campers for sheet")
        
        # Build CSV-like data for Google Sheets
        # Headers
        headers = ['First Name', 'Last Name', 'Address', 'Town', 'Zip', 'Session Type', 'AM Bus', 'PM Bus']
        
        # Sort by last name
        sheet_data.sort(key=lambda x: (x['last_name'], x['first_name']))
        
        # Convert to rows
        rows = [headers]
        for camper in sheet_data:
            rows.append([
                camper['first_name'],
                camper['last_name'],
                camper['address'],
                camper['town'],
                camper['zip'],
                camper['session'],
                camper['am_bus'],
                camper['pm_bus']
            ])
        
        # Use the webhook URL to update the sheet
        webhook_url = os.environ.get('GOOGLE_SHEETS_WEBHOOK_URL', '')
        
        if webhook_url:
            # Try to use webhook for update
            logger.info("Attempting update via webhook...")
            
            async with httpx.AsyncClient(timeout=60.0) as client:
                try:
                    response = await client.post(
                        webhook_url,
                        json={
                            'action': 'updateOutputSheet',
                            'sheetId': OUTPUT_SHEET_ID,
                            'data': rows
                        }
                    )
                    
                    if response.status_code == 200:
                        logger.info("✓ Sheet updated via webhook")
                        return {
                            "status": "success",
                            "message": f"Updated {len(sheet_data)} campers in Google Sheet",
                            "sheet_url": f"https://docs.google.com/spreadsheets/d/{OUTPUT_SHEET_ID}/edit",
                            "rows_written": len(rows),
                            "method": "webhook"
                        }
                except Exception as e:
                    logger.warning(f"Webhook update failed: {str(e)}, trying direct API...")
        
        # If no webhook or webhook failed, try direct API access
        # This requires a service account with access to the sheet
        try:
            from google.oauth2 import service_account
            from googleapiclient.discovery import build
            
            # Check for service account credentials
            creds_path = os.environ.get('GOOGLE_APPLICATION_CREDENTIALS', '/app/backend/service-account.json')
            
            if os.path.exists(creds_path):
                credentials = service_account.Credentials.from_service_account_file(
                    creds_path,
                    scopes=['https://www.googleapis.com/auth/spreadsheets']
                )
                
                service = build('sheets', 'v4', credentials=credentials)
                
                # Clear existing data
                logger.info("Clearing existing data...")
                service.spreadsheets().values().clear(
                    spreadsheetId=OUTPUT_SHEET_ID,
                    range='Sheet1!A1:H1000'
                ).execute()
                
                # Write new data
                logger.info(f"Writing {len(rows)} rows...")
                result = service.spreadsheets().values().update(
                    spreadsheetId=OUTPUT_SHEET_ID,
                    range='Sheet1!A1',
                    valueInputOption='USER_ENTERED',
                    body={'values': rows}
                ).execute()
                
                logger.info(f"✓ Updated {result.get('updatedCells', 0)} cells")
                
                return {
                    "status": "success",
                    "message": f"Updated {len(sheet_data)} campers in Google Sheet",
                    "sheet_url": f"https://docs.google.com/spreadsheets/d/{OUTPUT_SHEET_ID}/edit",
                    "rows_written": len(rows),
                    "cells_updated": result.get('updatedCells', 0),
                    "method": "direct_api"
                }
            else:
                logger.warning("No service account credentials found")
                
        except ImportError:
            logger.warning("Google API client not fully configured")
        except Exception as e:
            logger.error(f"Direct API update failed: {str(e)}")
        
        # Return data for manual update if automated methods fail
        return {
            "status": "manual_required",
            "message": "Automated update not available. Use the data below to update manually.",
            "sheet_url": f"https://docs.google.com/spreadsheets/d/{OUTPUT_SHEET_ID}/edit",
            "total_campers": len(sheet_data),
            "headers": headers,
            "sample_data": rows[:10],
            "full_data_available": True,
            "instructions": "Copy the data from /api/export-campers-csv endpoint to update the sheet manually"
        }
        
    except Exception as e:
        logger.error(f"Error updating output sheet: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/export-campers-csv")
async def export_campers_csv():
    """Export all campers as CSV for manual sheet update"""
    from fastapi.responses import StreamingResponse
    import io
    import csv
    
    # Get all campers
    all_campers = await db.campers.find({}).to_list(None)
    
    # Prepare CSV
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Headers
    writer.writerow(['First Name', 'Last Name', 'Address', 'Town', 'Zip', 'Session Type', 'AM Bus', 'PM Bus'])
    
    seen = set()
    for camper in sorted(all_campers, key=lambda x: (x.get('last_name', ''), x.get('first_name', ''))):
        camper_id = camper.get('_id', '')
        if camper_id.endswith('_PM'):
            continue
        
        name = f"{camper.get('first_name', '')} {camper.get('last_name', '')}"
        if name in seen:
            continue
        seen.add(name)
        
        am_bus = camper.get('am_bus_number', '')
        pm_bus = camper.get('pm_bus_number', '')
        if am_bus == 'NONE':
            am_bus = ''
        if pm_bus == 'NONE':
            pm_bus = ''
        
        writer.writerow([
            camper.get('first_name', ''),
            camper.get('last_name', ''),
            camper.get('location', {}).get('address', ''),
            camper.get('town', ''),
            camper.get('zip_code', ''),
            camper.get('session', camper.get('pickup_type', '')),
            am_bus,
            pm_bus
        ])
    
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=camper_bus_assignments.csv"}
    )


@api_router.get("/route-sheet/{bus_number}")
async def get_route_sheet(bus_number: str):
    """Get printable route sheet with turn-by-turn directions for a specific bus"""
    try:
        # Get campers for this bus (check both AM and PM bus fields)
        campers = await db.campers.find({
            "$or": [
                {"am_bus_number": bus_number},
                {"pm_bus_number": bus_number}
            ]
        }).to_list(None)
        
        if not campers:
            raise HTTPException(status_code=404, detail=f"No campers found for {bus_number}")
        
        # Generate route sheet with directions
        route_sheet = route_printer.generate_route_sheet(bus_number, campers)
        
        return route_sheet
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error generating route sheet: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/route-sheet/{bus_number}/print")
async def get_printable_route_sheet(bus_number: str):
    """Get printable HTML route sheet"""
    try:
        from fastapi.responses import HTMLResponse
        
        # Get campers for this bus
        campers = await db.campers.find({
            "$or": [
                {"am_bus_number": bus_number},
                {"pm_bus_number": bus_number}
            ]
        }).to_list(None)
        
        if not campers:
            return HTMLResponse(content=f"<h1>No campers found for {bus_number}</h1>", status_code=404)
        
        # Generate route sheet
        route_sheet = route_printer.generate_route_sheet(bus_number, campers)
        
        # Generate HTML
        html = route_printer.generate_printable_html(route_sheet)
        
        return HTMLResponse(content=html)
    except Exception as e:
        logging.error(f"Error generating printable route: {str(e)}")
        return HTMLResponse(content=f"<h1>Error: {str(e)}</h1>", status_code=500)

@api_router.get("/campers/filter")
async def filter_campers(bus_number: str = None, session: str = None, pickup_type: str = None):
    """Filter campers by bus, session, or pickup type"""
    try:
        query = {"bus_number": {"$exists": True, "$nin": ["NONE", ""]}}
        
        if bus_number:
            query["bus_number"] = bus_number
        
        if session:
            query["session"] = {"$regex": session, "$options": "i"}
        
        if pickup_type:
            query["pickup_type"] = pickup_type
        
        campers = await db.campers.find(query, {"_id": 0}).to_list(None)
        
        return {
            "status": "success",
            "count": len(campers),
            "campers": campers
        }
    except Exception as e:
        logging.error(f"Error filtering campers: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/reports/missing-addresses")
async def get_missing_addresses_report():
    """Get report of campers with bus assignments but no addresses"""
    try:
        missing = await db.campers.find({
            "location.latitude": 0.0,
            "bus_number": {"$exists": True, "$ne": "NONE"}
        }).to_list(None)
        
        report = []
        for camper in missing:
            report.append({
                "first_name": camper.get('first_name'),
                "last_name": camper.get('last_name'),
                "bus_number": camper.get('bus_number'),
                "session": camper.get('session'),
                "town": camper.get('town'),
                "zip_code": camper.get('zip_code')
            })
        
        return {
            "status": "warning",
            "count": len(report),
            "message": f"{len(report)} campers need addresses to appear on map",
            "campers": report
        }
    except Exception as e:
        logging.error(f"Error generating report: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/campers/{camper_id}/change-bus")
async def change_camper_bus(camper_id: str, am_bus_number: str = None, pm_bus_number: str = None):
    """Manually override bus assignment - updates database AND Google Sheet instantly"""
    try:
        # Get camper data first
        camper = await db.campers.find_one({"_id": camper_id})
        if not camper:
            raise HTTPException(status_code=404, detail="Camper not found")
        
        updates = {}
        
        if am_bus_number:
            updates["am_bus_number"] = am_bus_number
            updates["bus_color"] = get_bus_color(am_bus_number)
        
        if pm_bus_number:
            updates["pm_bus_number"] = pm_bus_number
        
        if not updates:
            raise HTTPException(status_code=400, detail="No bus assignments provided")
        
        # Update database
        result = await db.campers.update_one(
            {"_id": camper_id},
            {"$set": updates}
        )
        
        if result.modified_count > 0:
            # INSTANTLY update Google Sheet via webhook (using GET with query params)
            # HARDCODED URL to avoid environment variable caching issues in production
            webhook_url = "https://script.google.com/macros/s/AKfycbw8JoFhHDgyigOLy8Y6jbKxC-dB-x_FivZHVTsI29fUzcRZmJ--dz3EmpVkTOEWXSkn/exec"
            print(f"=== WEBHOOK DEBUG ===")
            print(f"Using hardcoded webhook URL")
            
            try:
                async with httpx.AsyncClient(timeout=15.0, follow_redirects=True) as client:
                    # Always send BOTH AM and PM bus numbers to keep sheet in sync
                    am_bus_to_send = updates.get('am_bus_number') if updates.get('am_bus_number') else camper.get('am_bus_number', '')
                    pm_bus_to_send = updates.get('pm_bus_number') if updates.get('pm_bus_number') else camper.get('pm_bus_number', '')
                    
                    params = {
                        "action": "updateBus",
                        "first_name": camper.get('first_name', '').strip(),
                        "last_name": camper.get('last_name', '').strip(),
                        "am_bus_number": am_bus_to_send,
                        "pm_bus_number": pm_bus_to_send
                    }
                    
                    print(f"Sending webhook with params: {params}")
                    response = await client.get(webhook_url, params=params)
                    print(f"Webhook response: {response.status_code} - {response.text[:300]}")
                    
                    if response.status_code == 200:
                        logger.info(f"✓ Google Sheet updated for {camper.get('first_name')} {camper.get('last_name')}: {response.text[:200]}")
                    else:
                        logger.warning(f"Webhook response: {response.status_code} - {response.text[:200]}")
            except Exception as e:
                print(f"WEBHOOK ERROR: {str(e)}")
                logger.error(f"Webhook error: {str(e)}")
            
            return {
                "status": "success",
                "message": "Updated bus assignments and Google Sheet"
            }
        else:
            raise HTTPException(status_code=404, detail="Camper not found")
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error changing bus: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/sync-to-google-sheet")
async def sync_bus_assignments_to_sheet():
    """
    Sync all bus assignments back to the Google Sheet.
    Updates the AM Bus and PM Bus columns for all campers.
    
    Sheet ID: 1QX0BSUuG889BjOYsTji8kYwT3VomSRE1j2_ZtxLd65k
    """
    import httpx
    import csv
    from io import StringIO
    
    logger.info("=== SYNCING BUS ASSIGNMENTS TO GOOGLE SHEET ===")
    
    try:
        # Get all campers from database
        db_campers = await db.campers.find({}).to_list(None)
        logger.info(f"Found {len(db_campers)} campers in database")
        
        # Build a lookup by first_name + last_name
        bus_lookup = {}
        for camper in db_campers:
            first_name = camper.get('first_name', '').strip()
            last_name = camper.get('last_name', '').strip()
            camper_id = camper.get('_id', '')
            
            # Skip PM-specific entries
            if camper_id.endswith('_PM'):
                continue
            
            key = f"{first_name}|{last_name}".lower()
            
            am_bus = camper.get('am_bus_number', '')
            pm_bus = camper.get('pm_bus_number', '')
            
            # Clean up
            if am_bus == 'NONE':
                am_bus = ''
            if pm_bus == 'NONE':
                pm_bus = ''
            
            bus_lookup[key] = {
                'am_bus': am_bus,
                'pm_bus': pm_bus
            }
        
        logger.info(f"Built lookup for {len(bus_lookup)} unique campers")
        
        # Read current sheet data
        sheet_id = CAMPMINDER_SHEET_ID
        csv_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv"
        
        async with httpx.AsyncClient(timeout=60.0, follow_redirects=True) as client:
            response = await client.get(csv_url)
            original_csv = response.text
        
        # Parse CSV
        if original_csv.startswith('\ufeff'):
            original_csv = original_csv[1:]
        
        reader = csv.DictReader(StringIO(original_csv))
        fieldnames = reader.fieldnames
        
        # Find the AM Bus and PM Bus column names and indices
        am_bus_col = None
        pm_bus_col = None
        am_bus_idx = None
        pm_bus_idx = None
        
        for idx, col in enumerate(fieldnames):
            if 'AM Bus' in col and 'Trans' in col:
                am_bus_col = col
                am_bus_idx = idx + 1  # 1-based for Sheets
            elif 'PM Bus' in col and 'Trans' in col:
                pm_bus_col = col
                pm_bus_idx = idx + 1
        
        if not am_bus_col or not pm_bus_col:
            logger.error(f"Could not find AM/PM Bus columns in sheet. Columns: {fieldnames}")
            return {
                "status": "error",
                "message": "Could not find AM Bus or PM Bus columns in sheet",
                "available_columns": [c for c in fieldnames if 'bus' in c.lower() or 'trans' in c.lower()]
            }
        
        logger.info(f"AM Bus column: {am_bus_col} (index {am_bus_idx})")
        logger.info(f"PM Bus column: {pm_bus_col} (index {pm_bus_idx})")
        
        # Count updates needed
        updates_needed = []
        rows = list(csv.DictReader(StringIO(original_csv)))
        
        for row_idx, row in enumerate(rows):
            first_name = row.get('First Name', '').strip()
            last_name = row.get('Last Name', '').strip()
            
            if not first_name or not last_name:
                continue
            
            key = f"{first_name}|{last_name}".lower()
            
            if key in bus_lookup:
                db_am = bus_lookup[key]['am_bus']
                db_pm = bus_lookup[key]['pm_bus']
                sheet_am = row.get(am_bus_col, '').strip()
                sheet_pm = row.get(pm_bus_col, '').strip()
                
                if db_am and db_am != sheet_am:
                    updates_needed.append({
                        'row': row_idx + 2,  # +2 for header and 1-based index
                        'col': am_bus_idx,
                        'name': f"{first_name} {last_name}",
                        'type': 'AM',
                        'from': sheet_am or 'EMPTY',
                        'to': db_am
                    })
                
                if db_pm and db_pm != sheet_pm:
                    updates_needed.append({
                        'row': row_idx + 2,
                        'col': pm_bus_idx,
                        'name': f"{first_name} {last_name}",
                        'type': 'PM',
                        'from': sheet_pm or 'EMPTY',
                        'to': db_pm
                    })
        
        logger.info(f"Found {len(updates_needed)} updates needed")
        
        if not updates_needed:
            return {
                "status": "success",
                "message": "No updates needed - sheet is already in sync",
                "updates_count": 0
            }
        
        # Try to use webhook if available
        webhook_url = os.environ.get('GOOGLE_SHEETS_WEBHOOK_URL', '')
        if webhook_url:
            try:
                async with httpx.AsyncClient(timeout=60.0, follow_redirects=True) as client:
                    # Send updates in batches
                    batch_size = 50
                    success_count = 0
                    
                    for i in range(0, len(updates_needed), batch_size):
                        batch = updates_needed[i:i+batch_size]
                        
                        payload = {
                            'action': 'updateBusAssignments',
                            'updates': [
                                {
                                    'row': u['row'],
                                    'col': u['col'],
                                    'value': u['to']
                                }
                                for u in batch
                            ]
                        }
                        
                        response = await client.post(
                            webhook_url,
                            json=payload,
                            headers={'Content-Type': 'application/json'}
                        )
                        
                        if response.status_code == 200:
                            success_count += len(batch)
                    
                    if success_count > 0:
                        return {
                            "status": "success",
                            "message": f"Updated {success_count} bus assignments in Google Sheet",
                            "updates_count": success_count
                        }
            except Exception as e:
                logger.warning(f"Webhook update failed: {str(e)}")
        
        # Return update information for manual or script-based update
        return {
            "status": "updates_available",
            "message": f"Found {len(updates_needed)} bus assignments that need updating in the Google Sheet",
            "sheet_url": f"https://docs.google.com/spreadsheets/d/{sheet_id}/edit",
            "am_bus_column": am_bus_col,
            "am_bus_column_index": am_bus_idx,
            "pm_bus_column": pm_bus_col,
            "pm_bus_column_index": pm_bus_idx,
            "updates_needed": updates_needed[:100],
            "total_updates": len(updates_needed),
            "instructions": [
                "Option 1: Copy the Google Apps Script below to your sheet",
                "Option 2: Download CSV from /api/export-campers-csv and import",
                "Option 3: Manually update each row in the sheet"
            ]
        }
        
    except Exception as e:
        logger.error(f"Error syncing to sheet: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/detect-changes")
async def detect_bus_assignment_changes():
    """
    Detect changes in bus assignments between database and Google Sheet.
    Identifies:
    - AM bus added (was empty, now has value)
    - PM bus added (was empty, now has value)
    - AM bus removed (had value, now empty)
    - PM bus removed (had value, now empty)
    - AM bus changed (different value)
    - PM bus changed (different value)
    
    After detection, automatically syncs changes to Google Sheet.
    """
    import csv
    from io import StringIO
    
    logger.info("=== DETECTING BUS ASSIGNMENT CHANGES ===")
    
    try:
        # Load database state
        db_campers = await db.campers.find({}).to_list(None)
        
        # Build lookup from database
        db_lookup = {}
        for camper in db_campers:
            first_name = camper.get('first_name', '').strip()
            last_name = camper.get('last_name', '').strip()
            camper_id = camper.get('_id', '')
            
            if camper_id.endswith('_PM'):
                continue
            
            key = f"{first_name}|{last_name}".lower()
            
            am_bus = camper.get('am_bus_number', '')
            pm_bus = camper.get('pm_bus_number', '')
            
            if am_bus == 'NONE':
                am_bus = ''
            if pm_bus == 'NONE':
                pm_bus = ''
            
            db_lookup[key] = {
                'first_name': first_name,
                'last_name': last_name,
                'am_bus': am_bus,
                'pm_bus': pm_bus,
                'address': camper.get('location', {}).get('address', ''),
                'has_location': bool(camper.get('location', {}).get('latitude'))
            }
        
        # Load Google Sheet state
        sheet_id = CAMPMINDER_SHEET_ID
        csv_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv"
        
        async with httpx.AsyncClient(timeout=60.0, follow_redirects=True) as client:
            response = await client.get(csv_url)
            csv_content = response.text
        
        if csv_content.startswith('\ufeff'):
            csv_content = csv_content[1:]
        
        reader = csv.DictReader(StringIO(csv_content))
        
        # Find bus columns
        fieldnames = reader.fieldnames
        am_bus_col = None
        pm_bus_col = None
        
        for col in fieldnames:
            if 'AM Bus' in col and 'Trans' in col:
                am_bus_col = col
            elif 'PM Bus' in col and 'Trans' in col:
                pm_bus_col = col
        
        # Build sheet lookup
        sheet_lookup = {}
        for row in reader:
            first_name = row.get('First Name', '').strip()
            last_name = row.get('Last Name', '').strip()
            
            if not first_name or not last_name:
                continue
            
            key = f"{first_name}|{last_name}".lower()
            
            sheet_am = row.get(am_bus_col, '').strip() if am_bus_col else ''
            sheet_pm = row.get(pm_bus_col, '').strip() if pm_bus_col else ''
            
            # Normalize NONE values
            if sheet_am.upper() == 'NONE':
                sheet_am = ''
            if sheet_pm.upper() == 'NONE':
                sheet_pm = ''
            
            sheet_lookup[key] = {
                'first_name': first_name,
                'last_name': last_name,
                'am_bus': sheet_am,
                'pm_bus': sheet_pm
            }
        
        # Detect changes
        changes = []
        
        for key, db_data in db_lookup.items():
            sheet_data = sheet_lookup.get(key)
            
            if not sheet_data:
                continue
            
            first_name = db_data['first_name']
            last_name = db_data['last_name']
            full_name = f"{first_name} {last_name}"
            
            db_am = db_data['am_bus']
            db_pm = db_data['pm_bus']
            sheet_am = sheet_data['am_bus']
            sheet_pm = sheet_data['pm_bus']
            
            # Check AM changes
            had_am = bool(sheet_am and sheet_am.startswith('Bus'))
            has_am = bool(db_am and db_am.startswith('Bus'))
            
            if not had_am and has_am:
                changes.append({
                    'name': full_name,
                    'type': 'AM_ADDED',
                    'old_value': sheet_am or 'EMPTY',
                    'new_value': db_am,
                    'message': f"{full_name}: AM bus ADDED ({db_am})"
                })
                logger.info(f"✓ {full_name}: AM bus ADDED ({db_am})")
                
            elif had_am and not has_am:
                changes.append({
                    'name': full_name,
                    'type': 'AM_REMOVED',
                    'old_value': sheet_am,
                    'new_value': 'EMPTY',
                    'message': f"{full_name}: AM bus REMOVED (was {sheet_am})"
                })
                logger.info(f"✓ {full_name}: AM bus REMOVED (was {sheet_am})")
                
            elif had_am and has_am and sheet_am != db_am:
                changes.append({
                    'name': full_name,
                    'type': 'AM_CHANGED',
                    'old_value': sheet_am,
                    'new_value': db_am,
                    'message': f"{full_name}: AM bus CHANGED ({sheet_am} → {db_am})"
                })
                logger.info(f"✓ {full_name}: AM bus CHANGED ({sheet_am} → {db_am})")
            
            # Check PM changes
            had_pm = bool(sheet_pm and sheet_pm.startswith('Bus'))
            has_pm = bool(db_pm and db_pm.startswith('Bus'))
            
            if not had_pm and has_pm:
                changes.append({
                    'name': full_name,
                    'type': 'PM_ADDED',
                    'old_value': sheet_pm or 'EMPTY',
                    'new_value': db_pm,
                    'message': f"{full_name}: PM bus ADDED ({db_pm})"
                })
                logger.info(f"✓ {full_name}: PM bus ADDED ({db_pm})")
                
            elif had_pm and not has_pm:
                changes.append({
                    'name': full_name,
                    'type': 'PM_REMOVED',
                    'old_value': sheet_pm,
                    'new_value': 'EMPTY',
                    'message': f"{full_name}: PM bus REMOVED (was {sheet_pm})"
                })
                logger.info(f"✓ {full_name}: PM bus REMOVED (was {sheet_pm})")
                
            elif had_pm and has_pm and sheet_pm != db_pm:
                changes.append({
                    'name': full_name,
                    'type': 'PM_CHANGED',
                    'old_value': sheet_pm,
                    'new_value': db_pm,
                    'message': f"{full_name}: PM bus CHANGED ({sheet_pm} → {db_pm})"
                })
                logger.info(f"✓ {full_name}: PM bus CHANGED ({sheet_pm} → {db_pm})")
        
        # Categorize changes
        am_added = [c for c in changes if c['type'] == 'AM_ADDED']
        pm_added = [c for c in changes if c['type'] == 'PM_ADDED']
        am_removed = [c for c in changes if c['type'] == 'AM_REMOVED']
        pm_removed = [c for c in changes if c['type'] == 'PM_REMOVED']
        am_changed = [c for c in changes if c['type'] == 'AM_CHANGED']
        pm_changed = [c for c in changes if c['type'] == 'PM_CHANGED']
        
        logger.info(f"Total changes detected: {len(changes)}")
        
        # If changes exist, sync to sheet
        sync_result = None
        if changes:
            logger.info("Syncing changes to Google Sheet...")
            # Call the sync endpoint logic
            sync_response = await sync_bus_assignments_to_sheet()
            sync_result = sync_response
        
        return {
            "status": "success",
            "total_changes": len(changes),
            "summary": {
                "am_added": len(am_added),
                "pm_added": len(pm_added),
                "am_removed": len(am_removed),
                "pm_removed": len(pm_removed),
                "am_changed": len(am_changed),
                "pm_changed": len(pm_changed)
            },
            "changes": changes,
            "am_added_campers": [c['name'] for c in am_added],
            "pm_added_campers": [c['name'] for c in pm_added],
            "sync_result": sync_result
        }
        
    except Exception as e:
        logger.error(f"Error detecting changes: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/refresh-and-sync")
async def refresh_and_sync():
    """
    Complete workflow:
    1. Refresh data from Google Sheet
    2. Detect any bus assignment changes
    3. Auto-assign buses to unassigned campers
    4. Sync all changes back to Google Sheet
    5. Return summary of all operations
    """
    logger.info("=== REFRESH AND SYNC WORKFLOW ===")
    
    results = {
        "step1_refresh": None,
        "step2_detect": None,
        "step3_sync": None,
        "summary": None
    }
    
    try:
        # Step 1: Trigger sync from Google Sheet
        logger.info("Step 1: Refreshing data from Google Sheet...")
        # This loads data from sheet into database
        await auto_sync_campminder()
        results["step1_refresh"] = {"status": "success", "message": "Data refreshed from Google Sheet"}
        
        # Step 2: Detect changes
        logger.info("Step 2: Detecting bus assignment changes...")
        detect_result = await detect_bus_assignment_changes()
        results["step2_detect"] = detect_result
        
        # Step 3: Sync back to sheet
        logger.info("Step 3: Syncing changes back to Google Sheet...")
        sync_result = await sync_bus_assignments_to_sheet()
        results["step3_sync"] = sync_result
        
        # Summary
        results["summary"] = {
            "status": "success",
            "message": "Refresh and sync completed successfully",
            "changes_detected": detect_result.get("total_changes", 0),
            "changes_synced": sync_result.get("updates_count", 0) if isinstance(sync_result, dict) else 0
        }
        
        return results
        
    except Exception as e:
        logger.error(f"Error in refresh and sync: {str(e)}")
        results["summary"] = {"status": "error", "message": str(e)}
        return results


@api_router.get("/google-apps-script")
async def get_google_apps_script():
    """
    Returns the Google Apps Script code that should be deployed to the Google Sheet
    to enable instant bus assignment updates.
    
    Instructions:
    1. Open your Google Sheet
    2. Extensions > Apps Script
    3. Paste this code
    4. Deploy as Web App (execute as yourself, allow anyone)
    5. Copy the deployment URL to GOOGLE_SHEETS_WEBHOOK_URL in .env
    """
    
    script = '''
// Google Apps Script for Bus Route Management
// Deploy this as a Web App to enable instant updates from the bus routing system

function doPost(e) {
  try {
    var sheet = SpreadsheetApp.getActiveSpreadsheet().getActiveSheet();
    var data = JSON.parse(e.postData.contents);
    
    if (data.action === 'updateBusAssignments') {
      // Batch update bus assignments
      var updates = data.updates || [];
      var updatedCount = 0;
      
      for (var i = 0; i < updates.length; i++) {
        var update = updates[i];
        if (update.row && update.col && update.value) {
          sheet.getRange(update.row, update.col).setValue(update.value);
          updatedCount++;
        }
      }
      
      return ContentService.createTextOutput(JSON.stringify({
        success: true,
        message: 'Updated ' + updatedCount + ' cells',
        count: updatedCount
      })).setMimeType(ContentService.MimeType.JSON);
      
    } else if (data.action === 'updateSingleCamper') {
      // Update a single camper's bus assignment
      var firstName = data.first_name;
      var lastName = data.last_name;
      var amBus = data.am_bus_number || '';
      var pmBus = data.pm_bus_number || '';
      
      // Find the row with this camper
      var dataRange = sheet.getDataRange();
      var values = dataRange.getValues();
      var headers = values[0];
      
      // Find column indices
      var firstNameCol = headers.indexOf('First Name');
      var lastNameCol = headers.indexOf('Last Name');
      var amBusCol = -1;
      var pmBusCol = -1;
      
      for (var h = 0; h < headers.length; h++) {
        if (headers[h].toString().indexOf('AM Bus') > -1 && headers[h].toString().indexOf('Trans') > -1) {
          amBusCol = h;
        }
        if (headers[h].toString().indexOf('PM Bus') > -1 && headers[h].toString().indexOf('Trans') > -1) {
          pmBusCol = h;
        }
      }
      
      if (firstNameCol < 0 || lastNameCol < 0 || amBusCol < 0 || pmBusCol < 0) {
        return ContentService.createTextOutput(JSON.stringify({
          success: false,
          message: 'Could not find required columns'
        })).setMimeType(ContentService.MimeType.JSON);
      }
      
      // Find and update the camper row
      for (var row = 1; row < values.length; row++) {
        if (values[row][firstNameCol] === firstName && values[row][lastNameCol] === lastName) {
          if (amBus) {
            sheet.getRange(row + 1, amBusCol + 1).setValue(amBus);
          }
          if (pmBus) {
            sheet.getRange(row + 1, pmBusCol + 1).setValue(pmBus);
          }
          
          return ContentService.createTextOutput(JSON.stringify({
            success: true,
            message: 'Updated ' + firstName + ' ' + lastName,
            row: row + 1
          })).setMimeType(ContentService.MimeType.JSON);
        }
      }
      
      return ContentService.createTextOutput(JSON.stringify({
        success: false,
        message: 'Camper not found: ' + firstName + ' ' + lastName
      })).setMimeType(ContentService.MimeType.JSON);
    }
    
    return ContentService.createTextOutput(JSON.stringify({
      success: false,
      message: 'Unknown action'
    })).setMimeType(ContentService.MimeType.JSON);
    
  } catch (error) {
    return ContentService.createTextOutput(JSON.stringify({
      success: false,
      error: error.toString()
    })).setMimeType(ContentService.MimeType.JSON);
  }
}

function doGet(e) {
  return ContentService.createTextOutput(JSON.stringify({
    status: 'ok',
    message: 'Bus Route Sheet Webhook is running'
  })).setMimeType(ContentService.MimeType.JSON);
}

// Test function - run this to verify the script works
function testScript() {
  var sheet = SpreadsheetApp.getActiveSpreadsheet().getActiveSheet();
  var headers = sheet.getRange(1, 1, 1, sheet.getLastColumn()).getValues()[0];
  
  Logger.log('Headers found: ' + headers.join(', '));
  
  var amBusCol = -1;
  var pmBusCol = -1;
  
  for (var h = 0; h < headers.length; h++) {
    if (headers[h].toString().indexOf('AM Bus') > -1) {
      amBusCol = h + 1;
      Logger.log('AM Bus column: ' + amBusCol + ' (' + headers[h] + ')');
    }
    if (headers[h].toString().indexOf('PM Bus') > -1) {
      pmBusCol = h + 1;
      Logger.log('PM Bus column: ' + pmBusCol + ' (' + headers[h] + ')');
    }
  }
  
  Logger.log('Script is ready to use!');
}
'''
    
    return {
        "status": "success",
        "script": script,
        "instructions": [
            "1. Open your Google Sheet: https://docs.google.com/spreadsheets/d/1QX0BSUuG889BjOYsTji8kYwT3VomSRE1j2_ZtxLd65k/edit",
            "2. Go to Extensions > Apps Script",
            "3. Delete any existing code and paste the script above",
            "4. Save the project (Ctrl+S)",
            "5. Click 'Deploy' > 'New deployment'",
            "6. Select 'Web app' as the type",
            "7. Set 'Execute as' to 'Me'",
            "8. Set 'Who has access' to 'Anyone'",
            "9. Click 'Deploy' and authorize when prompted",
            "10. Copy the Web App URL",
            "11. Update GOOGLE_SHEETS_WEBHOOK_URL in the backend .env file with this URL"
        ]
    }


@api_router.get("/download/bus-assignments")
async def download_bus_assignments():
    """Download bus assignments as CSV with AM and PM bus columns"""
    from fastapi.responses import Response
    from io import StringIO
    import csv as csv_module
    
    try:
        campers = await db.campers.find({
            "$or": [
                {"am_bus_number": {"$exists": True, "$nin": ["NONE", ""]}},
                {"pm_bus_number": {"$exists": True, "$nin": ["NONE", ""]}}
            ]
        }).to_list(None)
        
        # Create CSV
        output = StringIO()
        writer = csv_module.writer(output)
        
        # Simple header - just name and bus numbers
        writer.writerow([
            'Last Name',
            'First Name', 
            'AM Bus',
            'PM Bus'
        ])
        
        # Track campers to avoid duplicates
        seen_campers = set()
        
        # Data rows - sorted by last name, first name
        for camper in sorted(campers, key=lambda x: (x.get('last_name', '').lower(), x.get('first_name', '').lower())):
            camper_id = camper.get('_id', '')
            
            # Skip _PM suffix entries
            if str(camper_id).endswith('_PM'):
                continue
            
            camper_key = f"{camper.get('last_name', '')}_{camper.get('first_name', '')}"
            if camper_key in seen_campers:
                continue
            seen_campers.add(camper_key)
            
            am_bus = camper.get('am_bus_number', '')
            pm_bus = camper.get('pm_bus_number', '')
            
            # Look for _PM entry for this camper (may have different PM bus)
            for c in campers:
                if str(c.get('_id', '')).endswith('_PM'):
                    if c.get('first_name') == camper.get('first_name') and c.get('last_name') == camper.get('last_name'):
                        pm_bus = c.get('pm_bus_number', pm_bus)
                        break
            
            # Display "NONE" as "N/A"
            if am_bus == 'NONE' or not am_bus:
                am_bus = 'N/A'
            if pm_bus == 'NONE' or not pm_bus:
                pm_bus = 'N/A'
            
            writer.writerow([
                camper.get('last_name', ''),
                camper.get('first_name', ''),
                am_bus,
                pm_bus
            ])
        
        output.seek(0)
        
        filename = f"bus_assignments_{datetime.now().strftime('%Y%m%d')}.csv"
        
        return Response(
            content=output.getvalue(),
            media_type="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
                "Content-Type": "text/csv; charset=utf-8",
                "Cache-Control": "no-cache, no-store, must-revalidate",
                "Pragma": "no-cache",
                "Expires": "0",
                "Access-Control-Expose-Headers": "Content-Disposition"
            }
        )
    except Exception as e:
        logging.error(f"Error generating download: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    global sync_task
    if sync_task:
        sync_task.cancel()
    client.close()

# Auto-sync functions
async def auto_sync_campminder():
    """Background task to automatically sync from Google Sheets - handles ADD, UPDATE, DELETE"""
    global last_sync_time
    
    logger.info("Starting auto-sync from CampMinder Google Sheet...")
    
    try:
        # Download CSV from Google Sheets
        sheet_id = CAMPMINDER_SHEET_ID
        if not sheet_id:
            logger.error("No CAMPMINDER_SHEET_ID configured")
            return
        
        csv_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv"
        
        async with httpx.AsyncClient(timeout=60.0, follow_redirects=True) as client:
            response = await client.get(csv_url)
            
            if response.status_code != 200:
                logger.error(f"Failed to download Google Sheet: {response.status_code}")
                return
            
            csv_content = response.text
            logger.info(f"Downloaded CSV from Google Sheets ({len(csv_content)} chars)")
        
        # Process CSV same as manual upload
        from io import StringIO
        import csv
        
        # Remove BOM if present
        if csv_content.startswith('\ufeff'):
            csv_content = csv_content[1:]
        
        csv_file = StringIO(csv_content)
        reader = csv.DictReader(csv_file)
        
        # Log column names for debugging
        print(f"CSV columns: {reader.fieldnames}")
        
        # Track camper IDs from sheet
        sheet_camper_ids = set()
        new_count = 0
        updated_count = 0
        row_count = 0
        
        for row in reader:
            row_count += 1
            am_method = row.get('Trans-AMDropOffMethod', '').strip()
            pm_method = row.get('Trans-PMDismissalMethod', '').strip()
            pm_bus_raw = row.get('2026Transportation M PM Bus', '').strip()
            
            # Debug log for any row with "Carrol" in last name
            last_name_check = row.get('Last Name', '').strip()
            first_name_check = row.get('First Name', '').strip()
            if 'carrol' in last_name_check.lower():
                print(f"FOUND CARROL ROW #{row_count}: {first_name_check} {last_name_check}")
                print(f"  AM Method: '{am_method}'")
                print(f"  PM Method: '{pm_method}'")
                print(f"  PM Bus Raw: '{pm_bus_raw}'")
            
            # Determine if camper needs AM bus based on transport method
            am_needs_bus = 'am bus' in am_method.lower()
            # Determine if camper needs PM bus based on transport method
            pm_needs_bus = 'pm bus' in pm_method.lower()
            
            # Debug for Carroll
            if 'carrol' in last_name_check.lower():
                print(f"  am_needs_bus={am_needs_bus}, pm_needs_bus={pm_needs_bus}")
            
            # Skip campers who don't need any bus transport
            if not am_needs_bus and not pm_needs_bus:
                if 'carrol' in last_name_check.lower():
                    print(f"  SKIPPING Charlie - no bus needed")
                continue
            
            # Get all required fields first
            first_name = row.get('First Name', '')
            last_name = row.get('Last Name', '')
            session = row.get('Enrolled Child Sessions', '')
            
            am_address = row.get('Trans-PickUpAddress', '')
            am_town = row.get('Trans-PickUpTown', '')
            am_zip = row.get('Trans-PickUpZip', '')
            
            pm_address = row.get('Trans-DropOffAddress', '')
            pm_town = row.get('Trans-DropOffTown', '')
            pm_zip = row.get('Trans-DropOffZip', '')
            
            am_bus = row.get('2026Transportation M AM Bus', '')
            pm_bus = row.get('2026Transportation M PM Bus', '')
            
            # Only use bus values if the transport method calls for bus
            # If method is Car Drop Off, After Care, Morning Care - force NONE
            if not am_needs_bus:
                am_bus = 'NONE'
            if not pm_needs_bus:
                pm_bus = 'NONE'
            
            # Check if this is a PM-only camper (needs PM bus but not AM bus)
            is_pm_only_camper = not am_needs_bus and pm_needs_bus
            
            # PRESERVE existing bus assignments, AUTO-ASSIGN if empty/NONE
            final_am_bus = None
            final_pm_bus = None
            
            if am_needs_bus:
                if am_bus and am_bus.strip() and 'NONE' not in am_bus.upper():
                    # Has valid AM bus in sheet - KEEP IT
                    final_am_bus = am_bus.strip()
                elif am_address.strip():
                    # Bus is empty/NONE but has address - AUTO-ASSIGN
                    if 'existing_routes' not in locals():
                        all_db_campers = await db.campers.find({"am_bus_number": {"$exists": True}}).to_list(None)
                        existing_routes = {}
                        for ec in all_db_campers:
                            bus_str = ec.get('am_bus_number', '')
                            if bus_str and 'NONE' not in bus_str.upper():
                                try:
                                    bus_num = int(''.join(filter(str.isdigit, bus_str)))
                                    if bus_num not in existing_routes:
                                        existing_routes[bus_num] = []
                                    if ec.get('location', {}).get('latitude', 0) != 0:
                                        existing_routes[bus_num].append({
                                            'lat': ec['location']['latitude'],
                                            'lng': ec['location']['longitude']
                                        })
                                except (ValueError, IndexError):
                                    pass
                    
                    location_temp = await geocode_address_cached(am_address, am_town, am_zip)
                    if location_temp:
                        optimal_bus = route_optimizer.find_optimal_bus(
                            {'lat': location_temp.latitude, 'lng': location_temp.longitude},
                            existing_routes
                        )
                        final_am_bus = f"Bus #{optimal_bus:02d}"
                        logger.info(f"AUTO-ASSIGNED (new): {first_name} {last_name} → {final_am_bus}")
                        
                        # SYNC AUTO-ASSIGNMENT BACK TO GOOGLE SHEET
                        # Trim names to match sheet lookup (handles trailing spaces)
                        try:
                            webhook_url = "https://script.google.com/macros/s/AKfycbw8JoFhHDgyigOLy8Y6jbKxC-dB-x_FivZHVTsI29fUzcRZmJ--dz3EmpVkTOEWXSkn/exec"
                            async with httpx.AsyncClient(timeout=15.0, follow_redirects=True) as webhook_client:
                                params = {
                                    "action": "updateBus",
                                    "first_name": first_name.strip(),
                                    "last_name": last_name.strip(),
                                    "am_bus_number": final_am_bus,
                                    "pm_bus_number": final_am_bus if pm_needs_bus else "NONE"
                                }
                                webhook_response = await webhook_client.get(webhook_url, params=params)
                                if webhook_response.status_code == 200:
                                    logger.info(f"✓ Auto-assignment synced to sheet: {first_name.strip()} {last_name.strip()} → {final_am_bus}")
                                else:
                                    logger.warning(f"Sheet sync failed for {first_name} {last_name}: {webhook_response.status_code}")
                        except Exception as we:
                            logger.warning(f"Failed to sync auto-assignment to sheet: {str(we)}")
            else:
                # AM not needed - set to NONE
                final_am_bus = "NONE"
            
            # Handle PM bus assignment based on transport method
            if pm_needs_bus:
                if pm_bus and pm_bus.strip() and 'NONE' not in pm_bus.upper():
                    # Has valid PM bus - KEEP IT
                    final_pm_bus = pm_bus.strip()
                elif final_am_bus and final_am_bus != "NONE":
                    # Use AM bus for PM if available
                    final_pm_bus = final_am_bus
                else:
                    # PM bus needed but no value - needs assignment
                    final_pm_bus = "NONE"
            else:
                # PM bus NOT needed (Car Drop Off, After Care, etc.) - set to NONE
                final_pm_bus = "NONE"
            
            # Filter out non-bus PM values
            if final_pm_bus and any(x in final_pm_bus.upper() for x in ['MAIN TENT', 'HOCKEY RINK', 'AUDITORIUM']):
                final_pm_bus = "NONE"
            
            # For PM-only campers, use PM address
            effective_address = am_address.strip() if am_needs_bus else pm_address.strip()
            effective_town = am_town.strip() if am_needs_bus else pm_town.strip()
            effective_zip = am_zip.strip() if am_needs_bus else pm_zip.strip()
            
            # Debug for Carroll
            if 'carrol' in last_name.lower():
                print(f"  AM Address: '{am_address}', PM Address: '{pm_address}'")
                print(f"  Effective Address: '{effective_address}'")
            
            # If no effective address, try the other one
            if not effective_address:
                effective_address = pm_address.strip() or am_address.strip()
                effective_town = pm_town.strip() or am_town.strip()
                effective_zip = pm_zip.strip() or am_zip.strip()
                if 'carrol' in last_name.lower():
                    print(f"  Fallback Effective Address: '{effective_address}'")
            
            # For campers without address but with valid bus, still add them (for route planning)
            has_any_bus = (final_am_bus and final_am_bus != "NONE") or (final_pm_bus and final_pm_bus != "NONE")
            
            if 'carrol' in last_name.lower():
                print(f"  final_am_bus={final_am_bus}, final_pm_bus={final_pm_bus}, has_any_bus={has_any_bus}")
            
            if not effective_address and not has_any_bus:
                # Skip campers with no address AND no bus
                if 'carrol' in last_name.lower():
                    print(f"  SKIPPING Charlie - no address and no bus")
                continue
            
            # Determine pickup type based on transport methods
            if am_needs_bus and pm_needs_bus:
                pickup_type_val = "AM & PM"
            elif am_needs_bus:
                pickup_type_val = "AM Pickup Only"
            elif pm_needs_bus:
                pickup_type_val = "PM Drop-off Only"
            else:
                pickup_type_val = "Unknown"
            
            # Calculate final PM values for campers with different addresses
            pm_final_address = pm_address if pm_address.strip() else am_address
            pm_final_town = pm_town if pm_town.strip() else am_town
            pm_final_zip = pm_zip if pm_zip.strip() else am_zip
            
            # Process camper with BOTH AM and PM info
            # Create ONE entry if same address, TWO entries if different addresses
            
            # Generate camper ID - use zip if available, otherwise use "NOADDR"
            id_zip = effective_zip if effective_zip else "NOADDR"
            camper_id = f"{last_name}_{first_name}_{id_zip}".replace(' ', '_')
            sheet_camper_ids.add(camper_id)
            
            if effective_address:
                location = await geocode_address_cached(effective_address, effective_town, effective_zip)
                if not location:
                    location = GeoLocation(latitude=0.0, longitude=0.0, address=f"GEOCODING FAILED: {effective_address}")
                    logger.warning(f"Geocoding failed: {first_name} {last_name} - {effective_address}")
                
                # Calculate offset based on existing campers at this address
                address_key = f"{location.latitude:.6f}_{location.longitude:.6f}"
                existing_at_address = await db.campers.count_documents({
                    "location.latitude": {"$gte": location.latitude - 0.001, "$lte": location.latitude + 0.001},
                    "location.longitude": {"$gte": location.longitude - 0.001, "$lte": location.longitude + 0.001}
                })
                
                offset = existing_at_address * 0.00002  # ~6 feet per existing camper
                
                # Determine bus color based on pickup type
                if pickup_type_val == "PM Drop-off Only":
                    bus_color = get_bus_color(final_pm_bus) if final_pm_bus != "NONE" else "#808080"
                elif final_am_bus == "NONE":
                    bus_color = "#808080"
                else:
                    bus_color = get_bus_color(final_am_bus)
                
                camper_doc = {
                    "_id": camper_id,
                    "first_name": first_name,
                    "last_name": last_name,
                    "session": session,
                    "location": {
                        "latitude": location.latitude + offset,
                        "longitude": location.longitude + offset,
                        "address": location.address
                    },
                    "town": effective_town,
                    "zip_code": effective_zip,
                    "pickup_type": pickup_type_val,
                    "am_bus_number": final_am_bus,
                    "pm_bus_number": final_pm_bus,
                    "bus_color": bus_color,
                    "created_at": datetime.now(timezone.utc)
                }
                
                result = await db.campers.replace_one({"_id": camper_id}, camper_doc, upsert=True)
                if result.upserted_id:
                    new_count += 1
                elif result.modified_count > 0:
                    updated_count += 1
            else:
                # No address - still add for route planning
                # For PM-only campers, use PM bus color
                if is_pm_only and final_pm_bus and final_pm_bus != "NONE":
                    bus_color = get_bus_color(final_pm_bus)
                    pickup_type_val = "PM Drop-off Only - NO ADDRESS"
                elif final_am_bus and final_am_bus != "NONE":
                    bus_color = get_bus_color(final_am_bus)
                    pickup_type_val = "NO ADDRESS"
                else:
                    bus_color = "#808080"
                    pickup_type_val = "NO ADDRESS - NO BUS"
                
                camper_doc = {
                    "_id": camper_id,
                    "first_name": first_name,
                    "last_name": last_name,
                    "session": session,
                    "location": {"latitude": 0.0, "longitude": 0.0, "address": "ADDRESS NEEDED"},
                    "town": effective_town or "UNKNOWN",
                    "zip_code": effective_zip or "UNKNOWN",
                    "pickup_type": pickup_type_val,
                    "am_bus_number": final_am_bus,
                    "pm_bus_number": final_pm_bus,
                    "bus_color": bus_color,
                    "created_at": datetime.now(timezone.utc)
                }
                result = await db.campers.replace_one({"_id": camper_id}, camper_doc, upsert=True)
                if result.upserted_id:
                    new_count += 1
            
            # Only create separate PM entry if:
            # 1. PM address is different from AM address
            # 2. AND camper needs both AM and PM bus (not PM-only, they already have their PM entry)
            # 3. AND PM bus is needed
            has_different_pm_address = pm_final_address.strip() and pm_final_address != am_address
            if has_different_pm_address and am_needs_bus and pm_needs_bus:
                camper_id_pm = f"{last_name}_{first_name}_{pm_zip}_PM".replace(' ', '_')
                sheet_camper_ids.add(camper_id_pm)
                
                pm_location = await geocode_address_cached(pm_final_address, pm_final_town, pm_final_zip)
                if not pm_location:
                    pm_location = GeoLocation(latitude=0.0, longitude=0.0, address=f"GEOCODING FAILED: {pm_final_address}")
                    logger.warning(f"PM geocoding failed: {first_name} {last_name} - {pm_final_address}")
                
                # Calculate offset for PM address based on existing campers
                existing_at_pm_address = await db.campers.count_documents({
                    "location.latitude": {"$gte": pm_location.latitude - 0.001, "$lte": pm_location.latitude + 0.001},
                    "location.longitude": {"$gte": pm_location.longitude - 0.001, "$lte": pm_location.longitude + 0.001}
                })
                
                pm_offset = existing_at_pm_address * 0.00008
                
                camper_doc_pm = {
                    "_id": camper_id_pm,
                    "first_name": first_name,
                    "last_name": last_name,
                    "session": session,
                    "location": {
                        "latitude": pm_location.latitude + pm_offset,
                        "longitude": pm_location.longitude + pm_offset,
                        "address": pm_location.address
                    },
                    "town": pm_final_town,
                    "zip_code": pm_final_zip,
                    "pickup_type": "PM Drop-off Only",
                    "am_bus_number": final_am_bus,
                    "pm_bus_number": final_pm_bus,
                    "bus_color": get_bus_color(final_pm_bus),
                    "created_at": datetime.now(timezone.utc)
                }
                result = await db.campers.replace_one({"_id": camper_id_pm}, camper_doc_pm, upsert=True)
                if result.upserted_id:
                    new_count += 1
                elif result.modified_count > 0:
                    updated_count += 1
        
        # Delete campers no longer in sheet
        all_db_campers = await db.campers.find({}).to_list(None)
        deleted_count = 0
        for db_camper in all_db_campers:
            if db_camper['_id'] not in sheet_camper_ids:
                await db.campers.delete_one({"_id": db_camper['_id']})
                deleted_count += 1
                logger.info(f"Deleted: {db_camper.get('first_name')} {db_camper.get('last_name')}")
        
        last_sync_time = datetime.now(timezone.utc)
        logger.info(f"Auto-sync complete: {new_count} new, {updated_count} updated, {deleted_count} deleted")
        
        # AUTOMATICALLY apply sibling offset after sync
        await apply_sibling_offset(db)
        
        await db.sync_status.update_one(
            {"_id": "auto_sync"},
            {"$set": {
                "last_sync": last_sync_time,
                "new_campers": new_count,
                "updated_campers": updated_count,
                "deleted_campers": deleted_count,
                "status": "success",
                "source": "Google Sheets"
            }},
            upsert=True
        )
        
    except Exception as e:
        logger.error(f"Error in auto-sync: {str(e)}")
        await db.sync_status.update_one(
            {"_id": "auto_sync"},
            {"$set": {
                "last_sync": datetime.now(timezone.utc),
                "status": "error",
                "error": str(e)
            }},
            upsert=True
        )

async def sync_loop():
    """Continuous sync loop with retry logic"""
    global last_sync_time, db_connected
    
    # Check if database is configured
    if db is None:
        logger.error("Database not configured - sync loop disabled")
        return
    
    # Wait longer on startup to let MongoDB Atlas connect (can be slow)
    logger.info("Waiting 30 seconds before first sync to allow MongoDB Atlas connection...")
    await asyncio.sleep(30)
    
    retry_count = 0
    max_retries = 5
    base_delay = 15  # Base delay in seconds
    
    while True:
        try:
            # Test database connection first with timeout
            logger.info("Testing database connection...")
            await asyncio.wait_for(db.command('ping'), timeout=30.0)
            db_connected = True
            logger.info("Database connection verified, starting sync...")
            await auto_sync_campminder()
            retry_count = 0  # Reset retry count on success
            logger.info("Sync completed successfully")
        except asyncio.TimeoutError:
            retry_count += 1
            db_connected = False
            logger.error(f"Database connection timeout (attempt {retry_count}/{max_retries})")
            
            if retry_count >= max_retries:
                logger.warning(f"Max retries reached, waiting 5 minutes before next attempt...")
                await asyncio.sleep(300)  # Wait 5 minutes
                retry_count = 0
            else:
                delay = base_delay * (2 ** retry_count)  # Exponential backoff
                logger.info(f"Retrying in {delay} seconds...")
                await asyncio.sleep(delay)
                continue
        except Exception as e:
            retry_count += 1
            logger.error(f"Error in sync loop (attempt {retry_count}/{max_retries}): {str(e)}")
            
            if retry_count >= max_retries:
                logger.warning(f"Max retries reached, waiting 5 minutes before next attempt...")
                await asyncio.sleep(300)  # Wait 5 minutes
                retry_count = 0
            else:
                delay = base_delay * (2 ** retry_count)  # Exponential backoff
                logger.info(f"Retrying in {delay} seconds...")
                await asyncio.sleep(delay)
                continue
        
        await asyncio.sleep(SYNC_INTERVAL_MINUTES * 60)

@asynccontextmanager
async def lifespan(app: FastAPI):
    global sync_task
    
    logger.info(f"Auto-sync enabled: {AUTO_SYNC_ENABLED}")
    if AUTO_SYNC_ENABLED:
        logger.info(f"Starting auto-sync from Google Sheets (interval: {SYNC_INTERVAL_MINUTES} min)")
        sync_task = asyncio.create_task(sync_loop())
    
    yield
    
    if sync_task:
        sync_task.cancel()
        try:
            await sync_task
        except asyncio.CancelledError:
            pass

# Recreate app with lifespan
app = FastAPI(lifespan=lifespan)
app.include_router(api_router)

# Health check endpoint for Kubernetes - MUST be after app recreation
@app.get("/health")
async def health_check():
    """Health check endpoint for Kubernetes liveness/readiness probes"""
    return {"status": "healthy"}

# Database status endpoint
@app.get("/db-status")
async def db_status():
    """Check database connection status"""
    try:
        # Try to ping the database
        await asyncio.wait_for(db.command('ping'), timeout=10.0)
        camper_count = await asyncio.wait_for(db.campers.count_documents({}), timeout=10.0)
        return {
            "status": "connected",
            "database": os.environ.get('DB_NAME', 'unknown'),
            "camper_count": camper_count,
            "mongo_url_type": "atlas" if ('mongodb.net' in os.environ.get('MONGO_URL', '') or 'mongodb+srv' in os.environ.get('MONGO_URL', '')) else "local"
        }
    except asyncio.TimeoutError:
        return {
            "status": "timeout",
            "error": "Database connection timed out after 10 seconds",
            "mongo_url_type": "atlas" if ('mongodb.net' in os.environ.get('MONGO_URL', '') or 'mongodb+srv' in os.environ.get('MONGO_URL', '')) else "local"
        }
    except Exception as e:
        return {
            "status": "error",
            "error": str(e),
            "error_type": type(e).__name__,
            "mongo_url_type": "atlas" if ('mongodb.net' in os.environ.get('MONGO_URL', '') or 'mongodb+srv' in os.environ.get('MONGO_URL', '')) else "local"
        }

# Force sync endpoint for production debugging
@app.post("/force-sync")
async def force_sync():
    """Force a sync from Google Sheets - useful for production debugging"""
    try:
        # First check DB connection
        await asyncio.wait_for(db.command('ping'), timeout=10.0)
        
        # Trigger sync
        await auto_sync_campminder()
        
        camper_count = await db.campers.count_documents({})
        return {
            "status": "success",
            "message": "Sync completed",
            "camper_count": camper_count
        }
    except asyncio.TimeoutError:
        return {
            "status": "error",
            "error": "Database connection timed out"
        }
    except Exception as e:
        return {
            "status": "error",
            "error": str(e),
            "error_type": type(e).__name__
        }

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

